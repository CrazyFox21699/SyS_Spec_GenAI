"""Export primary user-facing Excel workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FILL_APPROVED = PatternFill("solid", fgColor="C6EFCE")
FILL_BLOCKED = PatternFill("solid", fgColor="FFC7CE")
FILL_REVIEW = PatternFill("solid", fgColor="FFEB9C")
FILL_REJECTED = PatternFill("solid", fgColor="D9D9D9")


def _style_header(ws, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}1"


def _auto_width(ws, max_w: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_w)


def _row_fill(status: str) -> PatternFill | None:
    s = (status or "").lower()
    if "approved" in s:
        return FILL_APPROVED
    if "blocked" in s:
        return FILL_BLOCKED
    if "reject" in s:
        return FILL_REJECTED
    if "review" in s or "pending" in s:
        return FILL_REVIEW
    return None


def _write_sheet(ws, headers: list[str], rows: list[list[Any]], status_col: int | None = None) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_header(ws, len(headers))
    if status_col is not None:
        for ri in range(2, ws.max_row + 1):
            fill = _row_fill(str(ws.cell(ri, status_col).value or ""))
            if fill:
                for ci in range(1, len(headers) + 1):
                    ws.cell(ri, ci).fill = fill
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _auto_width(ws)


def export_generated_test_spec(output_dir: Path, bundle: dict[str, Any]) -> Path:
    path = output_dir / "generated_test_spec.xlsx"
    wb = Workbook()
    candidates = bundle.get("test_candidates") or []

    def cand_row(c: dict[str, Any]) -> list[Any]:
        op = c.get("operation") or {}
        exp = c.get("expectation") or []
        imp = c.get("improvement_suggestion") or {}
        return [
            c.get("id"),
            c.get("test_function"),
            c.get("event"),
            c.get("use_case_description"),
            str(c.get("precondition", ""))[:500],
            str(op.get("given", ""))[:500],
            str(op.get("when", ""))[:300],
            str(exp[0].get("description") if exp and isinstance(exp[0], dict) else exp)[:500],
            str(op.get("given", ""))[:200],
            str((c.get("traceability") or {}).get("source_evidence", ""))[:300],
            c.get("confidence"),
            c.get("review_status", "pending"),
            "",
            imp.get("suggested_description", "")[:300],
        ]

    headers = [
        "Test Case ID",
        "Test Function",
        "Event",
        "Use Case / Description",
        "Precondition",
        "Operation / Given",
        "When / Timing",
        "Expected Result",
        "Logic Path",
        "Source Evidence",
        "Confidence",
        "Review Status",
        "Issue Link",
        "Improvement Suggestion",
    ]

    ws = wb.active
    ws.title = "Test_Spec_Candidates"
    _write_sheet(ws, headers, [cand_row(c) for c in candidates], status_col=12)

    approved = [c for c in candidates if c.get("review_status") == "approved"]
    blocked = [c for c in candidates if c.get("status") == "blocked" or c.get("review_status") == "blocked"]

    ws2 = wb.create_sheet("Approved_Candidates")
    _write_sheet(ws2, headers, [cand_row(c) for c in approved], status_col=12)

    ws3 = wb.create_sheet("Blocked_Candidates")
    _write_sheet(ws3, headers, [cand_row(c) for c in blocked], status_col=12)

    ws4 = wb.create_sheet("Description_Improvements")
    imp_headers = [
        "Test Case ID",
        "Current Description",
        "Suggested Description",
        "Reason",
        "Missing Information",
        "Source Evidence",
        "Confidence",
        "Review Status",
    ]
    imp_rows = []
    for c in candidates:
        imp = c.get("improvement_suggestion")
        if not imp:
            continue
        imp_rows.append(
            [
                c.get("id"),
                c.get("use_case_description"),
                imp.get("suggested_description"),
                imp.get("reason"),
                ", ".join(imp.get("missing_information") or []),
                str(imp.get("source_evidence", "")),
                imp.get("confidence", "medium"),
                c.get("review_status", "pending"),
            ]
        )
    _write_sheet(ws4, imp_headers, imp_rows, status_col=8)

    wb.save(path)
    return path


def export_review_package(output_dir: Path, bundle: dict[str, Any]) -> Path:
    path = output_dir / "review_package.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    def add(name: str, headers: list[str], rows: list[list[Any]]) -> None:
        ws = wb.create_sheet(name[:31])
        _write_sheet(ws, headers, rows)

    classified = bundle.get("classified_files") or []
    add(
        "File_Classification",
        ["File", "Role", "Confidence", "Reasons", "User Confirm"],
        [
            [
                c.get("file", ""),
                c.get("role"),
                c.get("confidence"),
                "; ".join(c.get("reason") or []),
                c.get("user_confirmation_suggested"),
            ]
            for c in classified
        ],
    )

    signals = bundle.get("signals") or []
    add(
        "Extracted_Signals",
        ["Name", "Direction", "Source", "Confidence", "Review Required"],
        [
            [
                s.get("name"),
                s.get("direction"),
                str(s.get("source", "")),
                s.get("confidence"),
                s.get("review_required"),
            ]
            for s in signals
        ],
    )

    states = bundle.get("states") or []
    add(
        "Extracted_States",
        ["Name", "Mode", "Description", "Review Required"],
        [[st.get("name"), st.get("mode"), st.get("description"), st.get("review_required")] for st in states],
    )

    defs = bundle.get("condition_definitions") or []
    add(
        "Extracted_Conditions",
        ["Name", "Definition", "Source", "Review Required"],
        [[d.get("name"), d.get("definition"), str(d.get("source", "")), True] for d in defs],
    )

    trees = bundle.get("condition_trees") or []
    add(
        "Condition_Trees",
        ["ID", "Name", "Raw Condition", "Parse Status", "Source"],
        [
            [
                t.get("transition_id"),
                t.get("name"),
                str(t.get("raw_condition", ""))[:500],
                t.get("parse_status"),
                str(t.get("source", "")),
            ]
            for t in trees
        ],
    )

    timing = bundle.get("timing_constraints") or []
    add(
        "Timing_Constants",
        ["Raw", "Interpreted", "Confidence", "Review Required"],
        [
            [t.get("raw_text"), t.get("interpreted_as"), t.get("confidence"), t.get("review_required")]
            for t in timing
        ],
    )

    aliases = bundle.get("alias_map") or []
    add(
        "Alias_Map",
        ["Alias", "Target", "Raw", "Source"],
        [[a.get("alias"), a.get("target"), a.get("raw_text"), str(a.get("source", ""))] for a in aliases],
    )

    footnotes = bundle.get("footnote_definitions") or []
    add(
        "Footnote_Definitions",
        ["Ref", "Raw Text", "Definition", "Source", "Review Required"],
        [
            [
                f.get("ref"),
                f.get("raw_text"),
                f.get("definition"),
                str(f.get("source", "")),
                f.get("review_required"),
            ]
            for f in footnotes
        ],
    )

    transitions = bundle.get("transitions") or []
    add(
        "Diagram_Transitions",
        ["From", "To", "Event", "Condition", "Source"],
        [
            [
                t.get("from_state"),
                t.get("to_state"),
                t.get("event"),
                str(t.get("raw_condition", ""))[:300],
                str(t.get("source", "")),
            ]
            for t in transitions
        ],
    )

    semantics = bundle.get("diagram_semantics") or {}
    add(
        "Diagram_Semantics",
        ["From", "To", "Event", "Semantic Type", "Conditions", "Evidence", "Review Required"],
        [
            [
                row.get("from_state"),
                row.get("to_state"),
                row.get("event"),
                row.get("semantic_type"),
                "; ".join(row.get("conditions") or []),
                "; ".join(row.get("evidence_refs") or []),
                row.get("review_required"),
            ]
            for row in semantics.get("edges", [])
        ],
    )

    jp = bundle.get("japanese_interpretations") or []
    add(
        "LLM_Interpretation",
        ["File", "Snippet", "Interpretation"],
        [[j.get("file"), str(j.get("snippet", ""))[:200], j.get("interpretation", "")] for j in jp],
    )

    questions = bundle.get("review_questions") or []
    add("Review_Questions", ["#", "Question"], [[i + 1, q] for i, q in enumerate(questions)])

    two_col = bundle.get("two_column_tables") or []
    rows_tc = []
    for tbl in two_col:
        for r in tbl.get("rows") or []:
            rows_tc.append(
                [
                    tbl.get("table_id"),
                    r.get("row_no"),
                    r.get("control"),
                    r.get("condition_raw"),
                    r.get("indentation_level"),
                    r.get("detected_type"),
                    r.get("parsed_hint"),
                    str(r.get("source", "")),
                    r.get("issue_status"),
                ]
            )
    add(
        "Two_Column_Rows",
        ["Table ID", "Row", "Control", "Condition Raw", "Indent", "Type", "Parsed", "Source", "Issue"],
        rows_tc,
    )

    wb.save(path)
    return path


def export_logic_traceability(output_dir: Path, bundle: dict[str, Any]) -> Path:
    path = output_dir / "logic_traceability.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    matrix = bundle.get("traceability_matrix") or []
    headers = [
        "Trace ID",
        "Test Candidate ID",
        "Signal / Input",
        "Condition",
        "Parent Condition Group",
        "Logic Operator",
        "State Transition",
        "Output",
        "Constant / Timing",
        "Source Evidence",
        "Reason",
        "Confidence",
        "Review Status",
    ]
    ws = wb.create_sheet("Traceability_Chain")
    _write_sheet(
        ws,
        headers,
        [
            [
                r.get("trace_id"),
                r.get("test_candidate_id"),
                r.get("signal_input"),
                r.get("condition"),
                r.get("parent_condition_group"),
                r.get("logic_operator"),
                r.get("state_transition"),
                r.get("output"),
                r.get("constant_timing"),
                r.get("source_evidence"),
                r.get("reason"),
                r.get("confidence"),
                r.get("review_status"),
            ]
            for r in matrix
        ],
        status_col=13,
    )

    ast_rows = bundle.get("logic_ast_rows") or []
    ws2 = wb.create_sheet("Logic_AST")
    _write_sheet(
        ws2,
        [
            "Tree ID",
            "Parent Node ID",
            "Node ID",
            "Depth",
            "Node Type",
            "Operator",
            "Condition Name",
            "Raw Text",
            "Normalized Text",
            "Source",
            "Issue Status",
        ],
        [
            [
                r.get("tree_id"),
                r.get("parent_node_id"),
                r.get("node_id"),
                r.get("depth"),
                r.get("node_type"),
                r.get("operator"),
                r.get("condition_name"),
                r.get("raw_text"),
                r.get("normalized_text"),
                str(r.get("source", "")),
                r.get("issue_status"),
            ]
            for r in ast_rows
        ],
    )

    coverage = bundle.get("logic_path_coverage") or []
    ws3 = wb.create_sheet("Logic_Path_Coverage")
    _write_sheet(
        ws3,
        [
            "Test Candidate ID",
            "Covered Logic Path",
            "Positive / Negative",
            "OR Branch",
            "NOT Condition Covered",
            "Timing Covered",
            "Fallback Covered",
            "Missing Coverage",
            "Review Status",
        ],
        [
            [
                r.get("test_candidate_id"),
                r.get("covered_logic_path"),
                r.get("positive_negative"),
                r.get("or_branch"),
                r.get("not_condition_covered"),
                r.get("timing_covered"),
                r.get("fallback_covered"),
                r.get("missing_coverage"),
                r.get("review_status"),
            ]
            for r in coverage
        ],
        status_col=9,
    )

    wb.save(path)
    return path


def export_issue_list(output_dir: Path, bundle: dict[str, Any]) -> Path:
    path = output_dir / "issue_list.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues"
    issues = bundle.get("issues") or []
    headers = [
        "Issue ID",
        "Severity",
        "Type",
        "Raw Text",
        "Source File",
        "Source Location",
        "Reason",
        "Impact",
        "Affected Logic",
        "Affected Test Candidate",
        "Required Action",
        "Can Export",
        "Review Status",
    ]
    rows = []
    for i in issues:
        sev = i.get("severity", "")
        fill_row = [
            i.get("id"),
            sev,
            i.get("type"),
            i.get("message"),
            (i.get("source_ref") or {}).get("file", ""),
            str(i.get("source_ref", "")),
            i.get("message"),
            sev,
            ", ".join(i.get("affected_items") or [])[:200],
            "",
            i.get("required_action"),
            i.get("can_export", True),
            "pending",
        ]
        rows.append(fill_row)
    _write_sheet(ws, headers, rows, status_col=2)
    wb.save(path)
    return path


def export_all_excel(output_dir: Path, bundle: dict[str, Any]) -> dict[str, str]:
    return {
        "generated_test_spec": str(export_generated_test_spec(output_dir, bundle)),
        "review_package": str(export_review_package(output_dir, bundle)),
        "logic_traceability": str(export_logic_traceability(output_dir, bundle)),
        "issue_list": str(export_issue_list(output_dir, bundle)),
    }

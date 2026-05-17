from __future__ import annotations

from openpyxl import load_workbook

from src.exporters.customer_testspec_exporter import (
    build_customer_testspec_preview,
    derive_module_name,
    export_customer_testspec,
)


def test_derive_module_name_prefers_classified_system_spec() -> None:
    bundle = {
        "classified_files": [
            {"file": "/tmp/notes.md", "role": "other"},
            {"file": "/tmp/Body_Control_Module Spec.docx", "role": "system_spec"},
        ]
    }
    assert derive_module_name(bundle) == "Body_Control_Module_Spec"


def test_export_customer_testspec_uses_overlay_values(tmp_path) -> None:
    bundle = {
        "classified_files": [{"file": "/tmp/Power_Mode.docx", "role": "system_spec"}],
        "test_candidates": [
            {
                "id": "TC_PM_001",
                "test_function": "Power mode / state behavior",
                "event": "evaluate_shutdown",
                "use_case_description": "Base deterministic description",
                "precondition": [{"current_state": "RUN"}],
                "operation": {
                    "given": [{"signal": "MODE_STS", "value": 1}],
                    "when": [{"description": "Trigger shutdown judgment"}],
                },
                "expectation": [{"description": "VMODE_STS becomes 1"}],
                "traceability": {"source_evidence": ["spec.docx / table A / row 3"]},
                "confidence": "medium",
                "review_required": True,
                "review_status": "pending",
                "status": "candidate",
            }
        ],
        "ai_assists": {
            "candidate_overlays": {
                "TC_PM_001": {
                    "provider": "copilot_cli",
                    "logic_id": "LOGIC_001",
                    "control_name": "SHUTDOWN_DECISION",
                    "updated_at": "2026-05-16T00:00:00Z",
                    "changed_fields": ["UseCase", "ExpectedOutput"],
                    "evidence_refs": ["spec.docx / table A / row 3"],
                    "open_questions": ["Confirm debounce timing"],
                    "confidence": "medium",
                    "review_required": True,
                    "en": {
                        "use_case": "Verify shutdown decision while MODE_STS is asserted.",
                        "operation": "Evaluate shutdown judgment from RUN to OFF",
                        "expected_input": "Given: MODE_STS=1",
                        "expected_output": "Then: VMODE_STS=1",
                    },
                    "jp": {
                        "use_case": "MODE_STS が成立したときのシャットダウン判定を確認する。",
                        "operation": "RUN から OFF へのシャットダウン判定を評価する",
                        "expected_input": "Given: MODE_STS=1",
                        "expected_output": "Then: VMODE_STS=1",
                    },
                }
            }
        },
    }
    path = export_customer_testspec(tmp_path, bundle, language="JP")
    assert path.name == "TestSpec_Power_Mode_JP.xlsx"

    wb = load_workbook(path)
    assert wb.sheetnames == ["System Test Spec"]
    ws = wb["System Test Spec"]
    assert ws["D2"].value == "MODE_STS が成立したときのシャットダウン判定を確認する。"
    assert ws["E2"].value == "RUN から OFF へのシャットダウン判定を評価する"
    assert ws["F2"].value == "Given: MODE_STS=1"
    assert ws["G2"].value == "Then: VMODE_STS=1"
    assert ws["J2"].value == "copilot_cli"


def test_preview_binds_logic_transition_and_outputs_into_final_evidence() -> None:
    bundle = {
        "logic_blocks": [
            {
                "id": "LOGIC_010",
                "name": "SYS_ShutOff",
                "raw_expression": "(Mode_cmd = 1 OR IGN_SW = 0)",
                "source": {"file": "spec.xlsx", "sheet": "Logic", "row": 4},
            }
        ],
        "transitions": [
            {
                "id": "SM_XL_001",
                "from_state": "OFF",
                "to_state": "ON",
                "event": "Transition failed",
                "raw_condition": "Transition failed",
                "source": {"file": "spec.xlsx", "sheet": "Logic", "row": 20},
                "derivation": "excel_drawing_connector",
            }
        ],
        "state_rules": [
            {
                "name": "RELAY_1",
                "expression": "ON",
                "source": {"file": "spec.xlsx", "sheet": "Logic", "state": "ON"},
            }
        ],
        "diagram_semantics": {
            "edges": [
                {
                    "from_state": "OFF",
                    "to_state": "ON",
                    "event": "Transition failed",
                    "semantic_type": "explicit_arrow",
                    "evidence_refs": ["spec.xlsx / Logic / connector_1"],
                    "transition_ids": ["SM_XL_001"],
                }
            ]
        },
        "test_candidates": [
            {
                "id": "TC_PM_010",
                "test_function": "Power mode / state behavior",
                "event": "Transition failed",
                "use_case_description": "Positive path for transition",
                "precondition": [{"current_state": "OFF"}],
                "operation": {"given": [], "when": [{"description": "Trigger transition"}]},
                "expectation": [{"description": "Reach ON"}],
                "traceability": {"transition": "SM_XL_001", "source_evidence": ["spec.xlsx / row 20"]},
                "confidence": "medium",
                "review_required": True,
                "review_status": "pending",
                "status": "candidate",
            }
        ],
    }

    preview = build_customer_testspec_preview(bundle, language="EN")

    row = preview["rows"][0]
    binding = row["evidence_binding"]
    assert row["logic_id"] == "LOGIC_010"
    assert row["control_name"] == "SYS_ShutOff"
    assert "spec.xlsx / row 20" in row["source_evidence"]
    assert any(b.get("name") == "SYS_ShutOff" for b in binding.get("logic_blocks") or [])
    assert any(t.get("id") == "SM_XL_001" for t in binding.get("transitions") or [])
    assert any(e.get("from_state") == "OFF" for e in binding.get("diagram_edges") or [])
    assert any(o.get("name") == "RELAY_1" for o in binding.get("state_outputs") or [])


def test_preview_includes_engineer_and_attachment_definitions_in_expected_input() -> None:
    bundle = {
        "logic_blocks": [
            {
                "id": "LOGIC_100",
                "name": "SYS_SHUTOFF",
                "raw_expression": "(OK_SHUTOFF AND SAFE_ROUTE)",
                "source": {"file": "spec.xlsx", "sheet": "Logic", "row": 10},
            }
        ],
        "test_candidates": [
            {
                "id": "TC_PM_100",
                "test_function": "Power mode / logic",
                "event": "evaluate_SYS_SHUTOFF",
                "use_case_description": "Check shutdown logic",
                "operation": {"given": [], "when": []},
                "expectation": [],
                "traceability": {"logic_block": "LOGIC_100", "source_evidence": ["spec.xlsx / row 10"]},
                "confidence": "medium",
                "review_required": True,
                "review_status": "pending",
                "status": "candidate",
            }
        ],
        "ai_assists": {
            "engineer_definitions": {
                "OK_SHUTOFF": {"name": "OK_SHUTOFF", "definition": "MODE_STS = 1 and IGN_STS = 0", "logic_id": "LOGIC_100"}
            },
            "supplemental_definitions": {
                "LOGIC_100": [
                    {
                        "name": "SAFE_ROUTE",
                        "definition": "GEAR_POS = P and VEH_SPD = 0",
                        "source": {"file": "define.xlsx", "sheet": "Defs", "row": 4},
                    }
                ]
            },
        },
    }

    preview = build_customer_testspec_preview(bundle, language="EN")

    row = preview["rows"][0]
    assert "Given: MODE_STS=1 and IGN_STS = 0" in row["expected_input"]
    assert "Given: GEAR_POS=P" in row["expected_input"] or "VEH_SPD" in row["expected_input"]

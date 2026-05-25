from __future__ import annotations

from openpyxl import Workbook

from src.parsers.excel_parser import extract_excel_workbook


def test_extract_excel_workbook_parses_merged_event_condition_logic_table(tmp_path) -> None:
    path = tmp_path / "shutdown_logic.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Logic"

    ws["A1"] = "System will shut off when the"
    ws.merge_cells("A3:C3")
    ws["A3"] = "Event"
    ws.merge_cells("D3:E3")
    ws["D3"] = "Condition"

    ws.merge_cells("A4:A8")
    ws["A4"] = "SYS_ShutOff"
    ws.merge_cells("B4:B7")
    ws["B4"] = "AND"
    ws.merge_cells("C4:C5")
    ws["C4"] = "OR"

    ws.merge_cells("D4:E4")
    ws["D4"] = "Mode_cmd = 1"
    ws.merge_cells("D5:E5")
    ws["D5"] = "IGN_SW = 0"
    ws.merge_cells("D6:E6")
    ws["D6"] = "VehicleSpeed = 0"
    ws["C7"] = "NOT"
    ws.merge_cells("D7:E7")
    ws["D7"] = "Battery_OK = 1"
    ws["B8"] = "OR"
    ws.merge_cells("D8:E8")
    ws["D8"] = "DoorLock_STS = 1"

    wb.save(path)

    parsed = extract_excel_workbook(path, [])

    assert parsed["sheet_regions"]
    assert parsed["logic_blocks"]
    block = parsed["logic_blocks"][0]
    assert block["name"] == "SYS_ShutOff"
    assert block["parse_status"] == "ok"
    expr = block["raw_expression"]
    assert "Mode_cmd = 1" in expr
    assert "IGN_SW = 0" in expr
    assert "VehicleSpeed = 0" in expr
    assert "Battery_OK = 1" in expr
    assert "DoorLock_STS = 1" in expr


def test_extract_excel_workbook_parses_transition_region_with_actual_row_numbers(tmp_path) -> None:
    path = tmp_path / "state_transition.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "States"
    ws.append(["Previous State", "Next State", "Output"])
    ws.append(["OFF", "ON", "STSW = 1"])
    ws.append(["ON", "OFF", "OFF STS = 1"])
    wb.save(path)

    parsed = extract_excel_workbook(path, [])

    assert len(parsed["transition_candidates"]) == 2
    assert parsed["transition_candidates"][0]["from_state"] == "OFF"
    assert parsed["transition_candidates"][0]["to_state"] == "ON"
    assert parsed["transition_candidates"][0]["source"]["row"] == 2


def test_extract_excel_workbook_parses_sectioned_gate_spine_with_adjacent_cell_diagram(tmp_path) -> None:
    path = tmp_path / "sectioned_logic.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Spec"

    ws.merge_cells("A1:F1")
    ws["A1"] = "5. Control Conditions - Merged Logic Table"
    ws["A2"] = "Control"
    ws["B2"] = "Condition"
    ws["C2"] = "Definition / Detail"
    ws["D2"] = "Source Type"
    ws["E2"] = "Expected Parser Skill"
    ws["F2"] = "Review Note"
    ws["H1"] = "7. State Machine Diagram"
    ws["H4"] = "OFF\nPWR_STATE = 0\nRELAY_MAIN = OFF"
    ws["J4"] = "RUN\nPWR_STATE = 2\nRELAY_MAIN = ON"

    rows = [
        ["SYS_SHUTOFF", "AND", "root", "gate", "", ""],
        ["SYS_SHUTOFF", "PWR_REQ_VALID", "ref", "ref", "", ""],
        ["SYS_SHUTOFF", "VEHICLE_SAFE", "ref", "ref", "", ""],
        ["SYS_SHUTOFF", "OR", "branch", "gate", "", ""],
        ["SYS_SHUTOFF", "NORMAL_ROUTE", "ref", "ref", "", ""],
        ["SYS_SHUTOFF", "AND", "nested", "gate", "", ""],
        ["SYS_SHUTOFF", "BACKUP_ROUTE", "ref", "ref", "", ""],
        ["SYS_SHUTOFF", "T_SHUT_CONFIRM elapsed", "timer", "timing", "", ""],
        ["SYS_SHUTOFF", "NOT NOK_SHUTOFF", "negative", "gate", "", ""],
    ]
    for idx, row in enumerate(rows, start=3):
        for col, value in enumerate(row, start=1):
            ws.cell(row=idx, column=col).value = value

    wb.save(path)

    parsed = extract_excel_workbook(path, [])

    assert parsed["logic_blocks"]
    block = parsed["logic_blocks"][0]
    assert block["name"] == "SYS_SHUTOFF"
    assert block["parse_status"] == "ok"
    assert "NORMAL_ROUTE OR (BACKUP_ROUTE AND T_SHUT_CONFIRM elapsed)" in block["raw_expression"]
    assert parsed["diagram_meta"]

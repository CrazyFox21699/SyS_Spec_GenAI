from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.importers.customer_testspec_importer import import_customer_testspec_workbook, preview_testspec_workbook
from src.importers.job_bootstrap import bootstrap_from_bundle_dict, bootstrap_from_testspec_xlsx
from src.exporters.customer_testspec_exporter import CUSTOMER_TESTSPEC_HEADERS
from web.copilot_errors import classify_copilot_error


def _write_sample_testspec(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "System Test Spec"
    ws.append(CUSTOMER_TESTSPEC_HEADERS)
    ws.append(
        [
            1,
            "Garage mode",
            "GRMD_EIG",
            "Verify GRMD_EIG ON when in garage mode",
            "Given: GRMD_IFU=ON\nWhen: evaluate garage mode",
            "Given: GRMD_IFU=ON",
            "Then: GRMD_EIG=1",
            "",
            "",
            "",
            "",
            "medium",
            "review_required",
            "yes",
            "",
        ]
    )
    wb.save(path)


def _write_jp_sample_testspec(path: Path, *, with_title: bool = False, with_merged_rows: bool = False) -> None:
    """Team JP template matching real PowerMode TestSpec layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PowerModeON"
    if with_title:
        ws["B2"] = "Power Mode移行"
        header_row = 3
    else:
        header_row = 1
    for col_idx, header in enumerate(
        [
            "No",
            "機能テスト",
            "テストグループ",
            "イベント",
            "ユーザケース",
            "手順",
            "入力に対する期待値",
            "出力に対する期待値",
            "備考",
        ],
        start=1,
    ):
        ws.cell(header_row, col_idx, header)
    data_row = header_row + 1
    ws.cell(data_row, 1, 1)
    ws.cell(data_row, 2, "PowerModeON**")
    ws.cell(data_row, 3, "ADM1_ON state → ADM1_OFF state transition")
    ws.cell(data_row, 4, "Auto P response within T7")
    ws.cell(data_row, 5, "正常")
    ws.cell(
        data_row,
        6,
        "オートP要求信号に対する応答が制限時間T7[ms]以内に受信された場合、状態遷移が正常に行われること",
    )
    ws.cell(
        data_row,
        7,
        "Given: WMODE_CMD = 1 (手動運転)\nGiven: SP1 = 199 (V1 : 1.99 km/h、車速の境界値)\nWhen: T7[ms]/2時点で、APOK2 = 1 (オートP OK応答受信)",
    )
    ws.cell(
        data_row,
        8,
        "Then: V_PMODE_STS == 1 (ON)\nThen: ACCD == STD_ON\nWhen: T14[ms] (IGDオフ指令待ち時間)経過後\nThen: IGPD == STD_OFF",
    )
    ws.cell(data_row, 9, "-")
    if with_merged_rows:
        next_row = data_row + 1
        ws.cell(next_row, 1, 2)
        ws.cell(next_row, 6, "2行目 — merged grouping columns should carry forward")
        ws.cell(next_row, 7, "Given: SP1 = 200 (V2 : 2.00 km/h)")
        ws.cell(next_row, 8, "Then: V_PMODE_STS == 0 (OFF)")
        ws.cell(next_row, 9, "-")
    wb.save(path)


def test_import_jp_testspec_workbook(tmp_path: Path) -> None:
    xlsx = tmp_path / "TestSpec_PowerMode_JP.xlsx"
    _write_jp_sample_testspec(xlsx)
    prev = preview_testspec_workbook(xlsx)
    assert prev["ok"] is True
    imported = import_customer_testspec_workbook(xlsx, language="JP")
    assert len(imported["test_candidates"]) == 1
    cand = imported["test_candidates"][0]
    assert "PowerModeON" in cand["test_function"]
    overlay = imported["candidate_overlays"][cand["id"]]
    assert overlay["jp"]["use_case"] == "正常"
    assert "Given: WMODE_CMD" in overlay["jp"]["expected_input"]
    assert "Then: V_PMODE_STS" in overlay["jp"]["expected_output"]
    assert "オートP要求" in overlay["jp"]["operation"]
    assert len(cand["operation"]["given"]) >= 2
    assert len(cand["expectation"]) >= 2
    assert imported["export_language"] == "JP"


def test_import_jp_testspec_detects_section_title(tmp_path: Path) -> None:
    xlsx = tmp_path / "TestSpec_PowerMode_Title_JP.xlsx"
    _write_jp_sample_testspec(xlsx, with_title=True)
    imported = import_customer_testspec_workbook(xlsx, language="JP")
    layout = imported["testspec_layout"]
    assert layout["section_title"] == "Power Mode移行"
    assert layout["primary_sheet_name"] == "PowerModeON"
    assert layout["jp_template"] is True


def test_import_jp_testspec_fills_merged_group_columns(tmp_path: Path) -> None:
    xlsx = tmp_path / "TestSpec_PowerMode_Merged_JP.xlsx"
    _write_jp_sample_testspec(xlsx, with_merged_rows=True)
    imported = import_customer_testspec_workbook(xlsx, language="JP")
    assert len(imported["test_candidates"]) == 2
    second = imported["test_candidates"][1]
    assert second["test_function"] == "PowerModeON**"
    assert second["event"] == "Auto P response within T7"  # fill-down within same TF/TG block
    overlay = imported["candidate_overlays"][second["id"]]
    assert overlay["test_group"] == "ADM1_ON state → ADM1_OFF state transition"
    assert overlay["jp"]["use_case"] == "正常"


def test_import_export_roundtrip_preserves_jp_layout(tmp_path: Path) -> None:
    from src.exporters.customer_testspec_exporter import export_customer_testspec

    xlsx = tmp_path / "TestSpec_PowerMode_Roundtrip_JP.xlsx"
    _write_jp_sample_testspec(xlsx, with_title=True, with_merged_rows=True)
    bundle = bootstrap_from_testspec_xlsx(xlsx, language="JP")
    out = export_customer_testspec(tmp_path, bundle, language="JP")
    wb = __import__("openpyxl").load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    assert ws["B2"].value == "Power Mode移行"
    assert ws.title == "PowerModeON"


def test_import_testspec_workbook_creates_candidates(tmp_path: Path) -> None:
    xlsx = tmp_path / "TestSpec_Module_EN.xlsx"
    _write_sample_testspec(xlsx)
    imported = import_customer_testspec_workbook(xlsx, language="EN")
    assert len(imported["test_candidates"]) == 1
    assert imported["test_candidates"][0]["test_function"] == "Garage mode"
    assert imported["candidate_overlays"]
    assert len(imported["logic_blocks"]) == 1
    assert imported["logic_blocks"][0]["parse_status"] == "imported"


def test_bootstrap_from_testspec_xlsx(tmp_path: Path) -> None:
    xlsx = tmp_path / "TestSpec_Module_EN.xlsx"
    _write_sample_testspec(xlsx)
    bundle = bootstrap_from_testspec_xlsx(xlsx)
    assert bundle["bootstrap_source"] == "imported_testspec"
    assert len(bundle["test_candidates"]) == 1
    assert len(bundle["logic_blocks"]) >= 1
    assert bundle["summary"]["test_candidates"] == 1


def test_bootstrap_from_bundle_dict_adds_synthetic_logic() -> None:
    bundle = bootstrap_from_bundle_dict(
        {
            "test_candidates": [
                {
                    "id": "TC_001",
                    "test_function": "KCC signal",
                    "traceability": {"logic_id": "imported_KCC_signal", "control_name": "KCC signal"},
                    "operation": {"given": [], "when": []},
                    "expectation": [],
                    "status": "candidate",
                }
            ]
        },
        source="imported_yaml",
    )
    assert bundle["bootstrap_source"] == "imported_yaml"
    assert any(b.get("id") == "imported_KCC_signal" for b in bundle["logic_blocks"])


def test_import_jp_testspec_event_backfill_from_middle_row(tmp_path: Path) -> None:
    """Event in middle of a TF+TG block must backfill to rows before it (two-pass)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "EventBackfill"
    ws.append(["No", "機能テスト", "テストグループ", "イベント", "ユーザケース", "手順",
               "入力に対する期待値", "出力に対する期待値"])
    # Row 1: TF/TG set, Event BLANK
    ws.append([1, "FnA", "GroupX", "", "Normal", "op1", "Given: SIG=1", "Then: OUT=1"])
    # Row 2: TF/TG blank (visual group), Event present
    ws.append([2, "", "", "EventMid", "", "op2", "Given: SIG=2", "Then: OUT=2"])
    # Row 3: TF/TG blank, Event blank — should get EventMid too
    ws.append([3, "", "", "", "", "op3", "Given: SIG=3", "Then: OUT=3"])
    # Row 4: different TF — new block, must NOT get EventMid
    ws.append([4, "FnB", "GroupY", "OtherEvent", "Normal2", "op4", "Given: SIG=4", "Then: OUT=4"])
    xlsx = tmp_path / "EventBackfill.xlsx"
    wb.save(xlsx)
    imported = import_customer_testspec_workbook(xlsx, language="JP")
    cands = imported["test_candidates"]
    assert len(cands) == 4
    events = [c["event"] for c in cands]
    # Row 1 (before EventMid in block) must be backfilled
    assert events[0] == "EventMid", f"Row 1 should get backfilled Event, got: {events[0]!r}"
    # Row 2 has EventMid directly
    assert events[1] == "EventMid"
    # Row 3 (after EventMid, same block) should carry forward
    assert events[2] == "EventMid"
    # Row 4 is a different TF+TG block — must keep its own Event
    assert events[3] == "OtherEvent"


def test_import_jp_testspec_event_fallback_to_test_group(tmp_path: Path) -> None:
    """When no Event exists in any row of a TF+TG block, Event must be set to Test Group."""
    wb = Workbook()
    ws = wb.active
    ws.title = "EventFallback"
    ws.append(["No", "機能テスト", "テストグループ", "イベント", "ユーザケース", "手順",
               "入力に対する期待値", "出力に対する期待値"])
    # Block A: no Event at all — must fall back to "Evaluate NOK"
    ws.append([1, "FnA", "Evaluate NOK", "", "Normal", "op1", "Given: SIG=1", "Then: OUT=1"])
    ws.append([2, "", "", "", "", "op2", "Given: SIG=2", "Then: OUT=2"])
    # Block B: has an explicit Event — fallback must NOT override it
    ws.append([3, "FnB", "GroupY", "ExplicitEvent", "Normal2", "op3", "Given: SIG=3", "Then: OUT=3"])
    xlsx = tmp_path / "EventFallback.xlsx"
    wb.save(xlsx)
    imported = import_customer_testspec_workbook(xlsx, language="JP")
    cands = imported["test_candidates"]
    assert len(cands) == 3
    events = [c["event"] for c in cands]
    assert events[0] == "Evaluate NOK", f"Block A row 1 must get Test Group as fallback, got: {events[0]!r}"
    assert events[1] == "Evaluate NOK", f"Block A row 2 must get Test Group as fallback, got: {events[1]!r}"
    assert events[2] == "ExplicitEvent", f"Block B must keep its own Event, got: {events[2]!r}"


def test_copilot_error_taxonomy_no_context() -> None:
    err = classify_copilot_error(has_context_pack=False, has_logic_id=True)
    assert err["error_category"] == "no_context"
    assert "Build context" in err["error"]

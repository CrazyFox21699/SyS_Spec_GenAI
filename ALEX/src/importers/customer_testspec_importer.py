"""Reverse-import customer TestSpec workbooks into bundle candidates."""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.exporters.customer_testspec_exporter import CUSTOMER_TESTSPEC_HEADERS, CUSTOMER_TESTSPEC_JP_HEADERS
from src.importers.synthetic_logic import slug, synthetic_logic_block
from web.candidate_mutations import sanitize_id

# Canonical header → aliases (EN export + JP team template + variants)
_HEADER_ALIASES: dict[str, str] = {
    "no": "No",
    "no.": "No",
    "番号": "No",
    "test function": "Test Function",
    "testfunction": "Test Function",
    "機能テスト": "Test Function",
    "機能テスト(test function)": "Test Function",
    "test group": "Test Group",
    "testgroup": "Test Group",
    "テストグループ": "Test Group",
    "テストグループ(test group)": "Test Group",
    "event": "Event",
    "イベント": "Event",
    "イベント(event)": "Event",
    "usecase": "UseCase",
    "use case": "UseCase",
    "ユースケース": "UseCase",
    "ユースケース(use case)": "UseCase",
    "ユーザケース": "UseCase",
    "ユーザケース(user case)": "UseCase",
    "operation": "Operation",
    "手順": "Operation",
    "手順(operation)": "Operation",
    "expected value for input": "Expected value for input",
    "expected input": "Expected value for input",
    "入力に対する期待値": "Expected value for input",
    "入力に対する期待値(expected value for input)": "Expected value for input",
    "expected value for output": "Expected value for output",
    "expected output": "Expected value for output",
    "出力に対する期待値": "Expected value for output",
    "出力に対する期待値(expected value for output)": "Expected value for output",
    "remarks": "Remarks",
    "備考": "Remarks",
    "備考(remarks)": "Remarks",
    "candidate id": "Candidate ID",
    "source evidence": "Source Evidence",
    "ai provider": "AI Provider",
    "ai touched fields": "AI Touched Fields",
    "confidence": "Confidence",
    "review status": "Review Status",
    "engineer confirmation required": "Engineer Confirmation Required",
    "open questions": "Open Questions",
}

_CANONICAL_HEADERS = set(CUSTOMER_TESTSPEC_HEADERS) | {"Test Group", "Remarks"}

# Grouped JP template columns — forward-fill when Excel merged cells leave blanks.
_FILL_DOWN_COLUMNS = ("Test Function", "Test Group", "Event", "UseCase")

_IO_LINE_RE = re.compile(
    r"^(given|when|then|precondition)\s*[:：]\s*(.*)$",
    re.IGNORECASE,
)


def _norm_header(cell: Any) -> str:
    text = str(cell or "").strip()
    if not text:
        return ""
    cleaned = text.replace("\ufeff", "").replace("\u3000", " ").strip()
    key = cleaned.lower()
    if key in _HEADER_ALIASES:
        return _HEADER_ALIASES[key]
    first_line = cleaned.splitlines()[0].strip().lower()
    if first_line in _HEADER_ALIASES:
        return _HEADER_ALIASES[first_line]
    combined = re.sub(r"\s+", " ", cleaned).lower()
    if combined in _HEADER_ALIASES:
        return _HEADER_ALIASES[combined]
    return cleaned


def _header_map(header_row: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        canonical = _norm_header(cell)
        if canonical in _CANONICAL_HEADERS:
            mapping[canonical] = idx
    return mapping


def _detect_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    """Find header row within first 12 lines (title rows above headers are common)."""
    best_idx = 0
    best_map: dict[str, int] = {}
    best_score = -1
    for i, row in enumerate(rows[:12]):
        colmap = _header_map(list(row))
        score = len(colmap)
        has_fn = "Test Function" in colmap
        has_io = "Expected value for input" in colmap or "Expected value for output" in colmap
        if has_fn and has_io:
            score += 3
        if score > best_score:
            best_score = score
            best_idx = i
            best_map = colmap
    return best_idx, best_map


def _detect_section_title(rows: list[tuple[Any, ...]], header_idx: int) -> str:
    """Read section title from rows above the detected header (common JP layout)."""
    for row_idx in range(header_idx - 1, -1, -1):
        row = rows[row_idx]
        for col_idx in range(min(4, len(row))):
            text = str(row[col_idx] or "").strip()
            if not text or len(text) < 2:
                continue
            if _norm_header(text) in _CANONICAL_HEADERS:
                continue
            if text.lower() in {"no", "no."}:
                continue
            return text
    return ""


def _is_jp_template(colmap: dict[str, int]) -> bool:
    return "Test Group" in colmap or "Remarks" in colmap


def _fill_grouped_columns(row_list: list[Any], colmap: dict[str, int], carry: dict[str, str]) -> None:
    for key in _FILL_DOWN_COLUMNS:
        idx = colmap.get(key)
        if idx is None:
            continue
        raw = row_list[idx] if idx < len(row_list) else None
        text = str(raw).strip() if raw is not None else ""
        if text:
            carry[key] = text
        elif carry.get(key) and idx < len(row_list):
            row_list[idx] = carry[key]


def _cell(row: list[Any], colmap: dict[str, int], key: str) -> str:
    idx = colmap.get(key)
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def _split_io_lines(text: str) -> list[str]:
    lines: list[str] = []
    for raw in str(text or "").replace("\r\n", "\n").split("\n"):
        line = raw.strip()
        if line:
            lines.append(line)
    return lines


def _io_lines_by_kind(text: str, *, default_bucket: str = "given") -> dict[str, list[str]]:
    """Split workbook I/O cell into given/when/then line lists (verbatim)."""
    given: list[str] = []
    when: list[str] = []
    then: list[str] = []
    for line in _split_io_lines(text):
        m = _IO_LINE_RE.match(line)
        if m:
            kind = m.group(1).lower()
            if kind in ("given", "precondition"):
                given.append(line)
            elif kind == "when":
                when.append(line)
            elif kind == "then":
                then.append(line)
            continue
        if default_bucket == "then":
            then.append(line)
        else:
            given.append(line)
    return {"given": given, "when": when, "then": then}


def _extract_transition(lines: list[str]) -> dict[str, str]:
    from_state = ""
    to_state = ""
    for line in lines:
        if "transitioning" in line.lower() or "遷移" in line:
            continue
        m = re.search(r"(?:system\s+state|状態)\s*[=＝]\s*(\w+)", line, re.I)
        if m:
            if not from_state:
                from_state = m.group(1)
            else:
                to_state = m.group(1)
        m2 = re.search(r"(\w+)\s*→\s*(\w+)", line)
        if m2:
            from_state, to_state = m2.group(1), m2.group(2)
    out: dict[str, str] = {}
    if from_state:
        out["from"] = from_state
    if to_state:
        out["to"] = to_state
    return out


def _extract_timers(lines: list[str]) -> list[str]:
    found: list[str] = []
    for line in lines:
        for m in re.finditer(r"\b(T\d+(?:/\d+)?)\b", line, re.I):
            tok = m.group(1).upper()
            if tok not in found:
                found.append(tok)
    return found


def build_structured_io(
    *,
    no: str = "",
    operation: str = "",
    expected_input: str = "",
    expected_output: str = "",
    remarks: str = "",
) -> dict[str, Any]:
    """Canonical testcase model (CTM) for sync / regen / hash."""
    inp = _io_lines_by_kind(expected_input, default_bucket="given")
    out = _io_lines_by_kind(expected_output, default_bucket="then")
    when_all = inp["when"] + out["when"]
    then_all = out["then"]
    if not inp["given"] and not when_all and operation.strip():
        op = _io_lines_by_kind(operation, default_bucket="given")
        if op["given"] or op["when"]:
            inp = op
            when_all = op["when"]
    all_lines = inp["given"] + when_all + then_all
    return {
        "no": str(no or "").strip(),
        "operation": str(operation or "").strip(),
        "given_lines": inp["given"],
        "when_lines": when_all,
        "then_lines": then_all,
        "remarks": str(remarks or "").strip(),
        "transition": _extract_transition(all_lines),
        "timers": _extract_timers(all_lines),
    }


def compute_spec_hash(structured_io: dict[str, Any]) -> str:
    payload = {
        "no": structured_io.get("no"),
        "operation": structured_io.get("operation"),
        "given_lines": structured_io.get("given_lines") or [],
        "when_lines": structured_io.get("when_lines") or [],
        "then_lines": structured_io.get("then_lines") or [],
        "remarks": structured_io.get("remarks"),
    }
    raw = json.dumps(payload, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def compute_body_hash(structured_io: dict[str, Any]) -> str:
    """Hash I/O lines only — comment-only vs body stale detection."""
    payload = {
        "given_lines": structured_io.get("given_lines") or [],
        "when_lines": structured_io.get("when_lines") or [],
        "then_lines": structured_io.get("then_lines") or [],
    }
    raw = json.dumps(payload, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def structured_io_from_overlay(overlay: dict[str, Any], *, lang: str = "en") -> dict[str, Any]:
    if overlay.get("structured_io"):
        return dict(overlay["structured_io"])
    lang_key = "jp" if str(lang).upper().startswith("J") else "en"
    block = overlay.get(lang_key) or overlay.get("en") or overlay.get("jp") or {}
    return build_structured_io(
        no=str(overlay.get("no") or ""),
        operation=str(block.get("operation") or ""),
        expected_input=str(block.get("expected_input") or ""),
        expected_output=str(block.get("expected_output") or ""),
        remarks=str(overlay.get("remarks") or ""),
    )


def _parse_given_when(text: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    given: list[dict[str, Any]] = []
    when: list[dict[str, Any]] = []
    for line in _split_io_lines(text):
        m = _IO_LINE_RE.match(line)
        if m:
            kind = m.group(1).lower()
            body = m.group(2).strip() or line
            if kind in ("given", "precondition"):
                given.append({"note" if kind == "precondition" else "description": body})
            elif kind == "when":
                when.append({"description": body})
            continue
        given.append({"description": line})
    return given, when


def _parse_then_and_when(text: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse output column: Then: assertions + When: post-timing checks."""
    then_lines: list[dict[str, Any]] = []
    when_lines: list[dict[str, Any]] = []
    for line in _split_io_lines(text):
        m = _IO_LINE_RE.match(line)
        if m:
            kind = m.group(1).lower()
            body = m.group(2).strip() or line
            if kind == "then":
                then_lines.append({"description": body})
            elif kind == "when":
                when_lines.append({"description": body})
            continue
        then_lines.append({"description": line})
    return then_lines, when_lines


def _candidate_from_row(
    row: list[Any],
    colmap: dict[str, int],
    *,
    sheet_name: str,
    row_no: int,
    language: str,
    existing_ids: set[str],
) -> tuple[dict[str, Any], dict[str, Any], str]:
    test_function = _cell(row, colmap, "Test Function") or "Imported test"
    test_group = _cell(row, colmap, "Test Group")
    event = _cell(row, colmap, "Event") or test_group or "imported"
    use_case = _cell(row, colmap, "UseCase")
    operation_text = _cell(row, colmap, "Operation")
    expected_input = _cell(row, colmap, "Expected value for input")
    expected_output = _cell(row, colmap, "Expected value for output")
    remarks = _cell(row, colmap, "Remarks")
    candidate_id = _cell(row, colmap, "Candidate ID")
    review_status = _cell(row, colmap, "Review Status") or "review_required"
    confidence = _cell(row, colmap, "Confidence") or "medium"
    open_questions = _cell(row, colmap, "Open Questions")
    source_evidence = _cell(row, colmap, "Source Evidence")
    ai_provider = _cell(row, colmap, "AI Provider") or "imported_workbook"
    changed_fields_raw = _cell(row, colmap, "AI Touched Fields")
    changed_fields = [f.strip() for f in changed_fields_raw.split(",") if f.strip()]

    logic_key = slug(re.sub(r"\*+", "", test_function))
    logic_id = f"imported_{logic_key}"

    if candidate_id:
        try:
            cid = sanitize_id(candidate_id, field="candidate_id")
        except ValueError:
            cid = ""
    else:
        cid = ""
    if not cid or cid in existing_ids:
        n = 1
        while True:
            cid = f"TC_IMP_{n:03d}"
            if cid not in existing_ids:
                break
            n += 1
    existing_ids.add(cid)

    given, when_in = _parse_given_when(expected_input)
    expectation, when_out = _parse_then_and_when(expected_output)
    when = when_in + when_out

    if not given and not when_in and operation_text:
        op_given, op_when = _parse_given_when(operation_text)
        if op_given or op_when:
            given, when = op_given, op_when

    cand: dict[str, Any] = {
        "id": cid,
        "status": "candidate",
        "source": "imported_workbook",
        "test_function": test_function,
        "event": event,
        "use_case_description": use_case or test_group or operation_text[:200],
        "precondition": [],
        "operation": {"given": given, "when": when},
        "expectation": expectation,
        "traceability": {
            "logic_id": logic_id,
            "control_name": test_function,
            "source": "imported_workbook",
            "source_evidence": [source_evidence] if source_evidence else [f"{sheet_name} / row {row_no}"],
        },
        "why_recommended": ["Imported from existing TestSpec workbook"],
        "confidence": confidence,
        "review_required": "review" in review_status.lower() or review_status.lower() in {"", "pending"},
        "review_status": review_status or "review_required",
    }

    oq = [q.strip() for q in re.split(r"[;\n]", open_questions) if q.strip()]
    if remarks and remarks not in ("-", "—", "－"):
        oq.append(remarks)

    lang_key = "jp" if language.upper().startswith("J") else "en"
    overlay: dict[str, Any] = {
        "provider": ai_provider,
        "logic_id": logic_id,
        "control_name": test_function,
        "test_group": test_group,
        "changed_fields": changed_fields or ["UseCase", "Operation", "ExpectedInput", "ExpectedOutput"],
        "open_questions": oq,
        "confidence": confidence,
        "review_required": cand["review_required"],
        lang_key: {
            "use_case": use_case,
            "operation": operation_text,
            "expected_input": expected_input,
            "expected_output": expected_output,
        },
    }
    if lang_key == "en":
        overlay["jp"] = {"use_case": "", "operation": "", "expected_input": "", "expected_output": ""}
    else:
        overlay["en"] = {"use_case": "", "operation": "", "expected_input": "", "expected_output": ""}

    structured_io = build_structured_io(
        no=_cell(row, colmap, "No"),
        operation=operation_text,
        expected_input=expected_input,
        expected_output=expected_output,
        remarks=remarks,
    )
    overlay["structured_io"] = structured_io
    overlay["spec_hash"] = compute_spec_hash(structured_io)
    overlay["body_hash"] = compute_body_hash(structured_io)
    overlay["no"] = structured_io.get("no") or ""

    return cand, overlay, logic_id


_CUSTOMER_TESTSPEC_REQUIRED = ("Test Function",)
_IMPORT_IDENTITY_COLUMNS = ("Test Function", "UseCase", "Expected value for input", "Expected value for output", "Operation")


def _sheet_importable(colmap: dict[str, int]) -> bool:
    if "Test Function" not in colmap:
        return False
    return any(k in colmap for k in _IMPORT_IDENTITY_COLUMNS[1:])


def preview_testspec_workbook(path: Path) -> dict[str, Any]:
    """Check whether workbook headers match ALEX / team TestSpec layout."""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=12, values_only=True))
        if not rows:
            sheets.append({"name": sheet_name, "ok": False, "reason": "empty", "headers_found": []})
            continue
        header_idx, colmap = _detect_header_row(rows)
        raw_header = [str(c or "").strip() for c in rows[header_idx]]
        missing = [h for h in _CUSTOMER_TESTSPEC_REQUIRED if h not in colmap]
        ok = _sheet_importable(colmap)
        sheets.append(
            {
                "name": sheet_name,
                "ok": ok,
                "header_row": header_idx + 1,
                "headers_found": raw_header[:20],
                "mapped_columns": sorted(colmap.keys()),
                "missing_required": missing,
            }
        )
    wb.close()
    ok_sheets = [s for s in sheets if s.get("ok")]
    return {
        "ok": bool(ok_sheets),
        "sheets": sheets,
        "required_columns": list(_CUSTOMER_TESTSPEC_REQUIRED),
        "supported_columns": sorted(_CANONICAL_HEADERS),
        "hint": (
            "Import supports ALEX export headers (EN) and team JP templates "
            f"({', '.join(CUSTOMER_TESTSPEC_JP_HEADERS[1:4])}, …). "
            "Title rows above headers and merged grouping cells are detected automatically."
        ),
    }


def import_customer_testspec_workbook(
    path: Path,
    *,
    language: str = "EN",
    sheet_names: list[str] | None = None,
) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    targets = sheet_names or wb.sheetnames
    candidates: list[dict[str, Any]] = []
    overlays: dict[str, dict[str, Any]] = {}
    logic_groups: dict[str, str] = {}
    sheet_summary: list[dict[str, Any]] = []
    existing_ids: set[str] = set()
    layout_sheets: list[dict[str, Any]] = []
    section_title = ""
    primary_sheet_name = ""

    for sheet_name in targets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            sheet_summary.append({"name": sheet_name, "rows_imported": 0, "skipped": "empty"})
            continue
        header_idx, colmap = _detect_header_row(rows)
        if not _sheet_importable(colmap):
            sheet_summary.append(
                {
                    "name": sheet_name,
                    "rows_imported": 0,
                    "skipped": "header_mismatch",
                    "header_row": header_idx + 1,
                }
            )
            continue
        sheet_title = _detect_section_title(rows, header_idx)
        if sheet_title and not section_title:
            section_title = sheet_title
        if not primary_sheet_name:
            primary_sheet_name = sheet_name
        layout_sheets.append(
            {
                "name": sheet_name,
                "header_row": header_idx + 1,
                "section_title": sheet_title,
                "jp_template": _is_jp_template(colmap),
            }
        )
        imported = 0
        carry: dict[str, str] = {}
        for row_no, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
            if not any(str(c or "").strip() for c in row):
                continue
            row_list = list(row)
            _fill_grouped_columns(row_list, colmap, carry)
            if not (
                _cell(row_list, colmap, "Test Function")
                or _cell(row_list, colmap, "Expected value for input")
                or _cell(row_list, colmap, "Expected value for output")
            ):
                continue
            cand, overlay, logic_id = _candidate_from_row(
                row_list,
                colmap,
                sheet_name=sheet_name,
                row_no=row_no,
                language=language,
                existing_ids=existing_ids,
            )
            candidates.append(cand)
            overlays[cand["id"]] = overlay
            logic_groups[logic_id] = cand["test_function"]
            imported += 1
        sheet_summary.append({"name": sheet_name, "rows_imported": imported, "header_row": header_idx + 1})

    logic_blocks = [
        synthetic_logic_block(
            logic_id,
            control,
            source={"file": path.name, "sheet": sheet_name, "kind": "imported_testspec"},
        )
        for logic_id, control in logic_groups.items()
    ]

    return {
        "test_candidates": candidates,
        "candidate_overlays": overlays,
        "logic_blocks": logic_blocks,
        "sheet_summary": sheet_summary,
        "export_language": "JP" if language.upper().startswith("J") else "EN",
        "testspec_layout": {
            "section_title": section_title,
            "primary_sheet_name": primary_sheet_name,
            "header_row": layout_sheets[0]["header_row"] if layout_sheets else 1,
            "jp_template": any(s.get("jp_template") for s in layout_sheets),
            "sheets": layout_sheets,
        },
    }

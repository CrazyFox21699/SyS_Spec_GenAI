"""Customer-style single-sheet TestSpec workbook export."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.engine.concrete_test_values import (
    dedupe_expected_input_text,
    definition_to_given_line,
    expand_definition_to_given_lines,
    materialize_expected_input,
    materialize_expected_output,
)
from src.engine.evidence_binding import build_candidate_evidence_bindings
from src.engine.logic_compliance import check_logic_compliance
from src.engine.testspec_validator import validate_workbook_io
from src.models.evidence_model import format_locator
from src.utils.feature_flags import feature_enabled
from src.utils.config_path import get_config_path
from src.utils.yaml_utils import load_yaml

_CONFIG_PATH = get_config_path()

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FILL_READY = PatternFill("solid", fgColor="E2F0D9")
FILL_REVIEW = PatternFill("solid", fgColor="FFF2CC")
FILL_BLOCKED = PatternFill("solid", fgColor="FCE4D6")
CUSTOMER_TESTSPEC_HEADERS = [
    "No",
    "Test Function",
    "Event",
    "UseCase",
    "Operation",
    "Expected value for input",
    "Expected value for output",
    "Candidate ID",
    "Source Evidence",
    "AI Provider",
    "AI Touched Fields",
    "Confidence",
    "Review Status",
    "Engineer Confirmation Required",
    "Open Questions",
]


def _style_header(ws, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}1"


def _auto_width(ws, max_w: int = 42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_w)


def _row_fill(status: str) -> PatternFill | None:
    s = (status or "").lower()
    if "blocked" in s:
        return FILL_BLOCKED
    if "review" in s or "pending" in s:
        return FILL_REVIEW
    if "ready" in s or "approved" in s:
        return FILL_READY
    return None


def _write_sheet(ws, headers: list[str], rows: list[list[Any]], status_col: int | None = None) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_header(ws, len(headers))
    if status_col is not None:
        for ri in range(2, ws.max_row + 1):
            fill = _row_fill(str(ws.cell(ri, status_col).value or ""))
            if fill:
                for ci in range(1, len(headers) + 1):
                    ws.cell(ri, ci).fill = fill
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _auto_width(ws)


def derive_module_name(bundle: dict[str, Any]) -> str:
    files = bundle.get("classified_files") or []
    preferred = []
    for row in files:
        role = row.get("role")
        if role in ("system_spec", "behavior_logic", "test_reference"):
            preferred.append(row.get("file") or "")
    preferred.extend(row.get("file") or "" for row in files)
    for path in preferred:
        name = Path(str(path)).stem
        name = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")
        if name:
            return name[:48]
    return "Module"


def _format_precondition_line(item: Any) -> str:
    if isinstance(item, dict):
        if item.get("current_state"):
            return f"Precondition: System state = {item['current_state']}"
        if item.get("note"):
            note = str(item["note"]).strip()
            return note if note.lower().startswith("precondition:") else f"Precondition: {note}"
    text = str(item).strip()
    if text.startswith("{"):
        return ""
    return text if text.lower().startswith("precondition:") else f"Precondition: {text}"


def _format_given_line(item: Any) -> str:
    if isinstance(item, dict):
        sig = item.get("signal")
        val = item.get("value")
        note = item.get("note")
        if sig is not None and val is not None:
            return f"Given: {sig}={val}"
        if note:
            note = str(note).strip()
            return note if note.lower().startswith("given:") else f"Given: {note}"
        text = str(item).strip()
        return text if text.lower().startswith("given:") else f"Given: {text}"
    text = str(item).strip()
    return text if text.lower().startswith("given:") else f"Given: {text}"


def _format_when_line(item: Any) -> str:
    if isinstance(item, dict):
        if item.get("description"):
            desc = str(item["description"]).strip()
            return desc if desc.lower().startswith("when:") else f"When: {desc}"
        if item.get("timing"):
            timing = str(item["timing"]).strip()
            return timing if timing.lower().startswith("when:") else f"When: {timing}"
    text = str(item).strip()
    return text if text.lower().startswith("when:") else f"When: {text}"


def _format_then_line(item: Any) -> str:
    if isinstance(item, dict):
        sig = item.get("signal")
        val = item.get("value")
        if sig is not None and val is not None:
            return f"Then: {sig}={val}"
        desc = str(item.get("description") or item.get("review_note") or "").strip()
        if desc:
            if desc.lower().startswith("then:"):
                return desc
            become = re.match(r"^([A-Za-z][A-Za-z0-9_]*)\s+becomes\s+(.+)$", desc, flags=re.IGNORECASE)
            if become:
                return f"Then: {become.group(1)}={become.group(2).strip()}"
            return f"Then: {desc}"
        text = str(item).strip()
        return text if text.lower().startswith("then:") else f"Then: {text}"
    text = str(item).strip()
    return text if text.lower().startswith("then:") else f"Then: {text}"


def _join_preconditions(candidate: dict[str, Any]) -> str:
    lines = [_format_precondition_line(item) for item in candidate.get("precondition") or []]
    return "\n".join(line for line in lines if line)


def _join_given(candidate: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.extend(_format_precondition_line(item) for item in candidate.get("precondition") or [])
    for item in (candidate.get("operation") or {}).get("given") or []:
        line = _format_given_line(item)
        if line:
            lines.append(line)
    for item in (candidate.get("operation") or {}).get("when") or []:
        line = _format_when_line(item)
        if line:
            lines.append(line)
    return "\n".join(lines)


def _join_when(candidate: dict[str, Any]) -> str:
    lines = [_format_when_line(item) for item in (candidate.get("operation") or {}).get("when") or []]
    return "\n".join(line for line in lines if line)


def _join_expectation(candidate: dict[str, Any]) -> str:
    lines = [_format_then_line(item) for item in candidate.get("expectation") or []]
    return "\n".join(line for line in lines if line)


def _lang_text(overlay: dict[str, Any] | None, section: str, language: str) -> str:
    if not overlay:
        return ""
    lang = "jp" if language.upper() == "JP" else "en"
    return str((overlay.get(lang) or {}).get(section) or "")


def _overlay_changed(overlay: dict[str, Any] | None, label: str) -> bool:
    if not overlay:
        return False
    return label in set(overlay.get("changed_fields") or [])


def _workbook_text(
    overlay: dict[str, Any] | None,
    section: str,
    language: str,
    changed_label: str,
    fallback: str,
) -> str:
    """Prefer engineer-owned overlay; skip Python materialize when field was saved."""
    if _overlay_changed(overlay, changed_label):
        return _lang_text(overlay, section, language)
    text = _lang_text(overlay, section, language)
    if text:
        return text
    return fallback


_TERM_RE = re.compile(r"\b[A-Z][A-Z0-9_]+\b")
_IGNORE_TERMS = {"AND", "OR", "NOT", "TRUE", "FALSE", "OFF", "ON", "RUN", "ANY", "P"}


def _normalize_term(term: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(term or "").upper())


def _definition_lookup(bundle: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    lookup: dict[str, list[dict[str, Any]]] = {}

    def add_row(row: dict[str, Any], kind: str) -> None:
        name = str(row.get("name") or "").strip()
        definition = str(row.get("definition") or row.get("expression") or "").strip()
        if not name or not definition:
            return
        merged = {**row, "kind": kind, "definition": definition}
        for key in {name, _normalize_term(name)}:
            if not key:
                continue
            lookup.setdefault(key, [])
            if any(
                str(existing.get("name") or "") == name and str(existing.get("definition") or "") == definition
                for existing in lookup[key]
            ):
                continue
            lookup[key].append(merged)

    for row in bundle.get("condition_definitions") or []:
        add_row(row, "spec")
    ai = bundle.get("ai_assists") or {}
    for row in (ai.get("engineer_definitions") or {}).values():
        add_row(row, "engineer")
    for defs in (ai.get("supplemental_definitions") or {}).values():
        for row in defs or []:
            add_row(row, "attachment")
    return lookup


def _resolved_definition_lines(
    candidate: dict[str, Any],
    binding: dict[str, Any] | None,
    definition_lookup: dict[str, list[dict[str, Any]]],
) -> list[str]:
    texts: list[str] = []
    trace = candidate.get("traceability") or {}
    if trace.get("logic_block"):
        texts.append(str(trace.get("logic_block")))
    for block in (binding or {}).get("logic_blocks") or []:
        texts.append(str(block.get("raw_expression") or ""))
    for transition in (binding or {}).get("transitions") or []:
        texts.append(str(transition.get("raw_condition") or ""))
    operation = candidate.get("operation") or {}
    texts.extend(str(item.get("note") or "") for item in (operation.get("given") or []) if isinstance(item, dict))
    seen: set[str] = set()
    lines: list[str] = []
    for text in texts:
        for term in _TERM_RE.findall(text or ""):
            if term in _IGNORE_TERMS:
                continue
            if term in seen:
                continue
            rows = definition_lookup.get(term) or definition_lookup.get(_normalize_term(term)) or []
            if not rows:
                continue
            row = rows[0]
            seen.add(term)
            kind = str(row.get("kind") or "")
            defn = str(row.get("definition") or "")
            if kind == "engineer" and len(defn) > 48 and " " in defn and not defn.strip().startswith("="):
                continue
            line = definition_to_given_line(term, defn)
            if not line and re.search(r"\band\b", defn, re.I):
                lines.extend(expand_definition_to_given_lines(defn))
            elif line:
                lines.append(line)
    return lines


def _transition_operation(binding: dict[str, Any] | None) -> str:
    transitions = (binding or {}).get("transitions") or []
    if not transitions:
        return ""
    row = transitions[0]
    from_state = str(row.get("from_state") or "").strip()
    to_state = str(row.get("to_state") or "").strip()
    event = str(row.get("event") or "").strip()
    parts = []
    if from_state or to_state:
        parts.append(f"{from_state or '?'} -> {to_state or '?'}".strip())
    if event:
        parts.append(event)
    return " | ".join(part for part in parts if part)


def _state_output_lines(binding: dict[str, Any] | None) -> list[str]:
    lines = []
    for row in (binding or {}).get("state_outputs") or []:
        state = str(row.get("state") or "").strip()
        name = str(row.get("name") or "").strip()
        value = str(row.get("expression") or row.get("definition") or "").strip()
        if name and value:
            if state:
                lines.append(f"Then: {state} {name}={value}")
            else:
                lines.append(f"Then: {name}={value}")
    return lines


def _candidate_row(
    idx: int,
    candidate: dict[str, Any],
    overlay: dict[str, Any] | None,
    binding: dict[str, Any] | None,
    definition_lookup: dict[str, list[dict[str, Any]]],
    *,
    language: str,
) -> list[Any]:
    status = (overlay or {}).get("review_status_override") or candidate.get("review_status") or ("blocked" if candidate.get("status") == "blocked" else "review_required")
    use_case = _workbook_text(
        overlay,
        "use_case",
        language,
        "UseCase",
        str(candidate.get("use_case_description") or ""),
    )
    operation_fallback = _join_when(candidate) or _join_preconditions(candidate) or _transition_operation(binding)
    operation = _workbook_text(overlay, "operation", language, "Operation", operation_fallback)
    expected_input = _workbook_text(
        overlay,
        "expected_input",
        language,
        "ExpectedInput",
        materialize_expected_input(candidate, definition_lookup),
    )
    if not _overlay_changed(overlay, "ExpectedInput"):
        resolved_lines = _resolved_definition_lines(candidate, binding, definition_lookup)
        if resolved_lines:
            if expected_input:
                for line in resolved_lines:
                    if line not in expected_input:
                        expected_input = expected_input + "\n" + line
            else:
                expected_input = "\n".join(resolved_lines[:12])
    expected_input = dedupe_expected_input_text(expected_input)
    expected_output = _workbook_text(
        overlay,
        "expected_output",
        language,
        "ExpectedOutput",
        materialize_expected_output(candidate, binding),
    )
    trace = candidate.get("traceability") or {}
    evidence: list[str] = []
    for entry in trace.get("source_evidence") or []:
        if not entry:
            continue
        if isinstance(entry, dict):
            loc = format_locator(entry)
            if loc:
                evidence.append(loc)
            continue
        text = str(entry).strip()
        if text:
            evidence.append(text)
    return [
        idx,
        candidate.get("test_function", ""),
        candidate.get("event", ""),
        use_case,
        operation,
        expected_input,
        expected_output,
        candidate.get("id", ""),
        "; ".join(str(x) for x in evidence if x),
        overlay.get("provider", "") if overlay else "",
        ", ".join(overlay.get("changed_fields") or []) if overlay else "",
        overlay.get("confidence", candidate.get("confidence", "low")) if overlay else candidate.get("confidence", "low"),
        status,
        "yes" if (overlay or {}).get("review_required", candidate.get("review_required", True)) else "no",
        "; ".join((overlay or {}).get("open_questions") or []),
    ]


def _validator_enabled(bundle: dict[str, Any]) -> bool:
    if bundle.get("features_validator") is not None:
        return bool(bundle.get("features_validator"))
    cfg: dict[str, Any] = {}
    if _CONFIG_PATH.exists():
        try:
            cfg = load_yaml(_CONFIG_PATH)
        except OSError:
            cfg = {}
    return feature_enabled(cfg, "validator", default=False)


def build_customer_testspec_preview(
    bundle: dict[str, Any],
    *,
    language: str = "EN",
    validate_io: bool | None = None,
) -> dict[str, Any]:
    language = language.upper()
    ai = bundle.get("ai_assists") or {}
    overlays = ai.get("candidate_overlays") or {}
    candidates = [
        c
        for c in (bundle.get("test_candidates") or [])
        if str(c.get("status") or "candidate") != "removed"
    ]
    run_validator = _validator_enabled(bundle) if validate_io is None else bool(validate_io)
    bindings = build_candidate_evidence_bindings(bundle)
    definition_lookup = _definition_lookup(bundle)

    row_dicts: list[dict[str, Any]] = []
    for i, cand in enumerate(candidates, start=1):
        overlay = overlays.get(cand.get("id"))
        binding = bindings.get(str(cand.get("id") or ""))
        row = _candidate_row(i, cand, overlay, binding, definition_lookup, language=language)
        logic_id = (overlay or {}).get("logic_id") or (binding or {}).get("logic_id") or ""
        control_name = (overlay or {}).get("control_name") or (binding or {}).get("control_name") or ""
        if not logic_id:
            trace = str(cand.get("traceability") or "")
            logic_id = next(
                (
                    str(item.get("logic_id"))
                    for item in bundle.get("logic_review_items") or []
                    if str(item.get("logic_id", "")) in trace
                    or str(item.get("control_name") or "") in str(cand.get("event", ""))
                ),
                "",
            )
        entry: dict[str, Any] = {
            "no": row[0],
            "test_function": row[1],
            "event": row[2],
            "use_case": row[3],
            "operation": row[4],
            "expected_input": row[5],
            "expected_output": row[6],
            "candidate_id": row[7],
            "source_evidence": row[8],
            "ai_provider": row[9],
            "ai_touched_fields": row[10],
            "confidence": row[11],
            "review_status": row[12],
            "engineer_confirmation_required": row[13],
            "open_questions": row[14],
            "logic_id": logic_id,
            "control_name": control_name,
            "evidence_binding": binding or {},
            "source": cand.get("source", ""),
            "row": row,
        }
        if run_validator:
            validation = validate_workbook_io(row[5], row[6])
            logic = check_logic_compliance(cand, bundle, expected_input=row[5])
            validation["logic_compliance"] = logic
            if logic.get("misplaced_in_given"):
                validation["ok"] = False
                validation["issues"] = list(validation.get("issues") or []) + [
                    {
                        "code": "role_misplaced",
                        "severity": "error",
                        "message": (
                            "Output terms in Expected input (Given): "
                            + ", ".join(logic["misplaced_in_given"][:6])
                        ),
                    }
                ]
                validation["quality_score"] = min(int(validation.get("quality_score", 0)), 40)
            entry["validation"] = validation
            entry["logic_compliance"] = logic
        row_dicts.append(entry)

    summary_validation: dict[str, Any] | None = None
    if run_validator and row_dicts:
        scores = [int((r.get("validation") or {}).get("quality_score", 0)) for r in row_dicts]
        failed = sum(1 for r in row_dicts if not (r.get("validation") or {}).get("ok"))
        summary_validation = {
            "rows_total": len(row_dicts),
            "rows_failed": failed,
            "avg_quality_score": round(sum(scores) / len(scores), 1) if scores else 0,
        }

    return {
        "headers": list(CUSTOMER_TESTSPEC_HEADERS),
        "rows": row_dicts,
        "language": language,
        "validation_enabled": run_validator,
        "validation_summary": summary_validation,
    }


def export_customer_testspec(
    output_dir: Path,
    bundle: dict[str, Any],
    *,
    language: str = "EN",
) -> Path:
    language = language.upper()
    module_name = derive_module_name(bundle)
    filename = f"TestSpec_{module_name}_{language}.xlsx"
    path = output_dir / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "System Test Spec"

    preview = build_customer_testspec_preview(bundle, language=language)
    if preview.get("validation_enabled") and preview.get("validation_summary"):
        failed = int(preview["validation_summary"].get("rows_failed") or 0)
        cfg: dict[str, Any] = load_yaml(_CONFIG_PATH) if _CONFIG_PATH.exists() else {}
        export_cfg = cfg.get("export") if isinstance(cfg.get("export"), dict) else {}
        if failed and bool(export_cfg.get("strict")):
            raise ValueError(
                f"Export blocked: {failed} workbook row(s) failed I/O validation (export.strict=true)"
            )
    rows = [row["row"] for row in preview["rows"]]
    _write_sheet(ws, list(CUSTOMER_TESTSPEC_HEADERS), rows, status_col=13)

    wb.save(path)
    return path

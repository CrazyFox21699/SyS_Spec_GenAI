"""Excel workbook extraction with merged-cell fill and logic tables."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.engine.condition_tree_builder import parse_condition_tree
from src.engine.excel_priority_parser import annotate_logic_block_decision_mode
from src.engine.two_column_logic_parser import parse_table_to_logic_block
from src.parsers.excel_drawing_parser import extract_excel_drawing_semantics
from src.parsers.signal_table_parser import parse_signal_grid
from src.parsers.table_logic_parser import rows_from_grid, parse_transition_table
from src.models.evidence_model import make_evidence_ref
from src.parsers.two_column_table_parser import (
    ParsedTwoColumnTable,
    TwoColumnRow,
    parse_control_condition_grid,
    tables_to_dicts,
)
from src.utils.config_path import get_config_path
from src.utils.yaml_utils import load_yaml


def _excel_ingest_config() -> dict[str, Any]:
    try:
        cfg = load_yaml(get_config_path())
    except OSError:
        cfg = {}
    excel = cfg.get("excel") or {}
    return {
        "max_sheets": int(excel.get("max_sheets") or 20),
        "sheet_include_patterns": list(excel.get("sheet_include_patterns") or []),
    }


def _sheet_names_for_workbook(sheetnames: list[str]) -> list[str]:
    cfg = _excel_ingest_config()
    patterns = cfg.get("sheet_include_patterns") or []
    max_sheets = int(cfg.get("max_sheets") or 20)
    names = list(sheetnames)
    if patterns:
        compiled: list[re.Pattern[str]] = []
        for pat in patterns:
            try:
                compiled.append(re.compile(str(pat), re.IGNORECASE))
            except re.error:
                continue
        if compiled:
            names = [n for n in names if any(rx.search(n) for rx in compiled)]
    return names[:max_sheets]


def peek_excel_text(path: Path, max_chars: int = 8000) -> tuple[str, list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    sheet_names = list(wb.sheetnames)
    for name in sheet_names[:8]:
        ws = wb[name]
        rows = 0
        for row in ws.iter_rows(values_only=True):
            line = " | ".join("" if v is None else str(v) for v in row)
            if line.strip():
                parts.append(line)
            rows += 1
            if rows > 80:
                break
    wb.close()
    return "\n".join(parts)[:max_chars], sheet_names


def collect_merged_cell_evidence(
    ws: Worksheet,
    file_name: str,
    sheet_name: str,
) -> list[dict[str, Any]]:
    """Emit one EvidenceRef per merged range on the worksheet."""
    refs: list[dict[str, Any]] = []
    for mrange in ws.merged_cells.ranges:
        min_col, min_row, max_col_m, max_row_m = mrange.bounds
        val = ws.cell(row=min_row, column=min_col).value
        text = "" if val is None else str(val).strip()
        merge_range = f"R{min_row}C{min_col}:R{max_row_m}C{max_col_m}"
        locator = f"sheet {sheet_name} / {merge_range}"
        refs.append(
            make_evidence_ref(
                kind="table_merged_region",
                file=file_name,
                locator=locator,
                source={
                    "file": file_name,
                    "sheet": sheet_name,
                    "merge_range": merge_range,
                    "row_start": min_row,
                    "row_end": max_row_m,
                    "col_start": min_col,
                    "col_end": max_col_m,
                },
                excerpt=text,
                confidence="high" if text else "low",
                review_required=not bool(text),
            )
        )
    return refs


def collect_sheet_comments(
    ws: Worksheet,
    file_name: str,
    sheet_name: str,
    *,
    max_rows: int = 250,
    max_cols: int = 30,
) -> list[dict[str, Any]]:
    """Extract non-executable Excel cell review comments."""
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols):
        for cell in row:
            comment = getattr(cell, "comment", None)
            if not comment:
                continue
            text = str(getattr(comment, "text", "") or "").strip()
            if not text:
                continue
            out.append(
                {
                    "cell": cell.coordinate,
                    "text": text,
                    "author": str(getattr(comment, "author", "") or ""),
                    "executable": False,
                    "source": {
                        "file": file_name,
                        "sheet": sheet_name,
                        "row": cell.row,
                        "column": cell.column,
                    },
                }
            )
    return out


def _sheet_to_filled_rows(
    ws: Worksheet, max_row: int = 250, max_col: int = 30
) -> list[tuple[int, list[str]]]:
    """Read sheet with merged cells expanded (fill merged ranges with top-left value)."""
    merges = list(ws.merged_cells.ranges)
    merge_map: dict[tuple[int, int], str] = {}

    for mrange in merges:
        min_col, min_row, max_col_m, max_row_m = mrange.bounds
        val = ws.cell(row=min_row, column=min_col).value
        text = "" if val is None else str(val).strip()
        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                merge_map[(r, c)] = text

    rows: list[tuple[int, list[str]]] = []
    for r in range(1, min(ws.max_row or 1, max_row) + 1):
        row_cells: list[str] = []
        for c in range(1, min(ws.max_column or 1, max_col) + 1):
            if (r, c) in merge_map:
                row_cells.append(merge_map[(r, c)])
            else:
                v = ws.cell(row=r, column=c).value
                row_cells.append("" if v is None else str(v).strip())
        if any(row_cells):
            rows.append((r, row_cells))
    return rows


def _sheet_to_filled_grid(ws: Worksheet, max_row: int = 250, max_col: int = 30) -> list[list[str]]:
    return [cells for _, cells in _sheet_to_filled_rows(ws, max_row=max_row, max_col=max_col)]


def _find_nonempty_regions(
    rows: list[tuple[int, list[str]]],
    *,
    min_nonempty_cells: int = 4,
) -> list[dict[str, Any]]:
    occupied = {
        (ri, ci)
        for ri, (_, cells) in enumerate(rows)
        for ci, value in enumerate(cells)
        if str(value or "").strip()
    }
    seen: set[tuple[int, int]] = set()
    regions: list[dict[str, Any]] = []

    for start in sorted(occupied):
        if start in seen:
            continue
        stack = [start]
        cells_in_region: list[tuple[int, int]] = []
        seen.add(start)
        while stack:
            ri, ci = stack.pop()
            cells_in_region.append((ri, ci))
            for dr in (-1, 0, 1):
                for dc in (-1, 0, 1):
                    if dr == 0 and dc == 0:
                        continue
                    nxt = (ri + dr, ci + dc)
                    if nxt in occupied and nxt not in seen:
                        seen.add(nxt)
                        stack.append(nxt)
        if len(cells_in_region) < min_nonempty_cells:
            continue
        min_ri = min(ri for ri, _ in cells_in_region)
        max_ri = max(ri for ri, _ in cells_in_region)
        min_ci = min(ci for _, ci in cells_in_region)
        max_ci = max(ci for _, ci in cells_in_region)
        region_rows = [
            (actual_row, cells[min_ci : max_ci + 1])
            for actual_row, cells in rows[min_ri : max_ri + 1]
        ]
        regions.append(
            {
                "bbox": {
                    "row_start": region_rows[0][0],
                    "row_end": region_rows[-1][0],
                    "col_start": min_ci + 1,
                    "col_end": max_ci + 1,
                },
                "rows": region_rows,
                "nonempty_cells": len(cells_in_region),
            }
        )
    return regions


SECTION_HEADER_RE = re.compile(r"^\s*(\d+)\.\s+(.+?)\s*$")
STATE_CELL_TITLE_RE = re.compile(r"^[A-Z][A-Z0-9_ ]{1,30}$")


def _section_slices(rows: list[tuple[int, list[str]]]) -> list[dict[str, Any]]:
    starts: list[tuple[int, str]] = []
    for idx, (row_no, cells) in enumerate(rows):
        first = next((str(cell or "").strip() for cell in cells if str(cell or "").strip()), "")
        m = SECTION_HEADER_RE.match(first)
        if m:
            starts.append((idx, first))
    sections: list[dict[str, Any]] = []
    for i, (start_idx, title) in enumerate(starts):
        end_idx = starts[i + 1][0] if i + 1 < len(starts) else len(rows)
        section_rows = rows[start_idx:end_idx]
        sections.append({"title": title, "rows": section_rows})
    return sections


def _looks_like_event_condition_region(header: list[str]) -> bool:
    norm = [str(cell or "").strip().lower() for cell in header]
    has_event = any(token in {"event", "control", "item", "function"} for token in norm)
    has_condition = any("condition" in token or token == "cond" for token in norm)
    return has_event and has_condition


def _find_region_header_index(region_grid: list[list[str]]) -> int:
    for idx, row in enumerate(region_grid[:5]):
        norm = [str(cell or "").strip().lower() for cell in row]
        joined = " ".join(token for token in norm if token)
        if _looks_like_event_condition_region(row):
            return idx
        if "logic" in norm and any("condition" in token or token == "cond" for token in norm):
            return idx
        if "previous" in joined and "next" in joined:
            return idx
        if norm and "condition" in norm[0] and len(norm) > 1 and "definition" in norm[1]:
            return idx
        if norm and ("constant" in norm[0] or "parameter" in norm[0] or "name" in norm[0]) and len(norm) > 1 and ("value" in norm[1] or "definition" in norm[1]):
            return idx
    return 0


def _leading_header_span(cells: list[str]) -> int:
    started = False
    last = 0
    for idx, cell in enumerate(cells):
        text = str(cell or "").strip()
        if text:
            started = True
            last = idx + 1
        elif started:
            break
    return last or len(cells)


def _trim_section_rows(region_rows: list[tuple[int, list[str]]]) -> list[tuple[int, list[str]]]:
    if not region_rows:
        return []
    span = _leading_header_span(region_rows[0][1])
    if span <= 0:
        return region_rows
    return [(row_no, cells[:span]) for row_no, cells in region_rows]


def _combine_gate_parts(gate: str, parts: list[str]) -> str:
    values = [str(part or "").strip() for part in parts if str(part or "").strip()]
    if not values:
        return ""
    if gate == "NOT":
        if len(values) == 1:
            return f"NOT {values[0]}"
        return "NOT (" + " AND ".join(values) + ")"
    if len(values) == 1:
        return values[0]
    return "(" + f" {gate} ".join(values) + ")"


def _build_gate_spine_expression(tokens: list[str]) -> tuple[str, str]:
    root_gate = ""
    root_parts: list[str] = []
    groups: list[dict[str, Any]] = []

    def close_to(level: int) -> None:
        nonlocal groups, root_parts
        while len(groups) > level:
            grp = groups.pop()
            expr = _combine_gate_parts(str(grp.get("gate") or "AND"), list(grp.get("parts") or []))
            if not expr:
                continue
            if groups:
                groups[-1].setdefault("parts", []).append(expr)
            else:
                root_parts.append(expr)

    for raw_token in tokens:
        token = str(raw_token or "").strip()
        if not token:
            continue
        upper = token.upper()
        if upper in {"AND", "OR", "NOT"}:
            if not root_gate:
                root_gate = upper
                continue
            if upper == root_gate:
                if groups and str(groups[-1].get("gate") or "") != root_gate:
                    groups.append({"gate": upper, "parts": []})
                else:
                    close_to(0)
                continue
            groups.append({"gate": upper, "parts": []})
            continue
        if groups and len(groups[-1].get("parts") or []) >= 2:
            close_to(0)
        if groups:
            groups[-1].setdefault("parts", []).append(token)
        else:
            root_parts.append(token)

    close_to(0)
    return _combine_gate_parts(root_gate or "AND", root_parts), root_gate or "AND"


def _build_gate_spine_paths(tokens: list[str]) -> list[list[str]]:
    expr, root_gate = _build_gate_spine_expression(tokens)
    if not expr:
        return []
    parts = []
    root_parts: list[str] = []
    groups: list[dict[str, Any]] = []

    def close_to(level: int) -> None:
        nonlocal groups, root_parts
        while len(groups) > level:
            grp = groups.pop()
            subexpr = _combine_gate_parts(str(grp.get("gate") or "AND"), list(grp.get("parts") or []))
            if not subexpr:
                continue
            if groups:
                groups[-1].setdefault("parts", []).append(subexpr)
            else:
                root_parts.append(subexpr)

    for raw_token in tokens:
        token = str(raw_token or "").strip()
        if not token:
            continue
        upper = token.upper()
        if upper in {"AND", "OR", "NOT"}:
            if not root_parts and not groups and upper == root_gate:
                continue
            if upper == root_gate:
                if groups and str(groups[-1].get("gate") or "") != root_gate:
                    groups.append({"gate": upper, "parts": []})
                else:
                    close_to(0)
                continue
            groups.append({"gate": upper, "parts": []})
            continue
        if groups and len(groups[-1].get("parts") or []) >= 2:
            close_to(0)
        if groups:
            groups[-1].setdefault("parts", []).append(token)
        else:
            root_parts.append(token)

    close_to(0)
    for part in root_parts:
        parts.append([root_gate, part])
    return parts


def _parse_gate_spine_region(
    region_rows: list[tuple[int, list[str]]],
    source: dict[str, Any],
    *,
    table_id: str,
) -> list[dict[str, Any]]:
    if len(region_rows) < 2:
        return []
    trimmed_rows = _trim_section_rows(region_rows)
    header = [str(cell or "").strip().lower() for cell in trimmed_rows[0][1]]
    try:
        control_idx = next(
            idx for idx, token in enumerate(header) if token in {"event", "control", "item", "function"}
        )
        condition_idx = next(
            idx for idx, token in enumerate(header) if "condition" in token or token == "cond"
        )
    except StopIteration:
        return []
    detail_idx = next((idx for idx, token in enumerate(header) if "definition" in token or token == "detail"), None)

    grouped: dict[str, list[dict[str, Any]]] = {}
    order: list[str] = []
    current_control = ""
    for actual_row, cells in trimmed_rows[1:]:
        values = [(cell or "").strip() for cell in cells]
        if control_idx < len(values) and values[control_idx]:
            current_control = values[control_idx]
        if not current_control or condition_idx >= len(values):
            continue
        token = values[condition_idx]
        if not token:
            continue
        row = {
            "row_no": actual_row,
            "token": token,
            "detail": values[detail_idx] if detail_idx is not None and detail_idx < len(values) else "",
            "source": {**source, "row": actual_row},
        }
        if current_control not in grouped:
            grouped[current_control] = []
            order.append(current_control)
        grouped[current_control].append(row)

    blocks: list[dict[str, Any]] = []
    for index, control_name in enumerate(order, start=1):
        rows = grouped[control_name]
        tokens = [str(row.get("token") or "").strip() for row in rows if str(row.get("token") or "").strip()]
        expression, _ = _build_gate_spine_expression(tokens)
        if not expression:
            continue
        from src.engine.gate_spine_ast import build_gate_spine_ast

        tree = build_gate_spine_ast(rows)
        if str((tree or {}).get("type") or "") in ("empty", ""):
            tree = parse_condition_tree(expression)
        parse_status = "ok"
        parser_notes: list[dict[str, Any]] = []
        if str((tree or {}).get("parse_status") or "") != "ok":
            parser_notes.append(
                {
                    "severity": "info",
                    "parser_reason": "Expression recovered from Excel gate spine; inner timing/value clauses may still need review.",
                }
            )
        context_texts = tokens + [str(row.get("detail") or "") for row in rows]
        block = {
            "id": f"TC2_{table_id}_{index:02d}",
            "name": control_name,
            "raw_expression": expression,
            "tree": tree,
            "block_type": "two_column_control",
            "parse_status": parse_status,
            "review_required": True,
            "can_generate_candidates": True,
            "source": {**source, "control": control_name, "table_id": f"{table_id}_{index:02d}"},
            "table_kind": "logic",
            "row_paths": _build_gate_spine_paths(tokens),
            "issues": [],
            "parser_notes": parser_notes,
            "unresolved_refs": [],
        }
        annotate_logic_block_decision_mode(block, context_texts)
        blocks.append(block)
    return blocks


def _parse_event_condition_logic_region(
    region_rows: list[tuple[int, list[str]]],
    source: dict[str, Any],
    *,
    table_id: str,
) -> list[dict[str, Any]]:
    if len(region_rows) < 2:
        return []
    trimmed_rows = _trim_section_rows(region_rows)
    header = [str(cell or "").strip().lower() for cell in trimmed_rows[0][1]]
    try:
        control_idx = next(
            idx for idx, token in enumerate(header) if token in {"event", "control", "item", "function"}
        )
        condition_start = next(
            idx for idx, token in enumerate(header) if "condition" in token or token == "cond"
        )
    except StopIteration:
        return []
    if condition_start == control_idx + 1:
        return _parse_gate_spine_region(trimmed_rows, source, table_id=table_id)

    by_control: dict[str, list[TwoColumnRow]] = {}
    order: list[str] = []
    current_control = ""
    for actual_row, raw_cells in trimmed_rows[1:]:
        cells = [(cell or "").strip() for cell in raw_cells]
        if control_idx < len(cells) and cells[control_idx]:
            current_control = cells[control_idx]
        if not current_control:
            continue
        logic_tokens: list[str] = []
        for idx in range(control_idx + 1, min(condition_start, len(cells))):
            token = cells[idx].strip()
            upper = token.upper()
            if upper in {"AND", "OR", "NOT"}:
                logic_tokens.append(upper)
        condition_parts: list[str] = []
        for idx in range(condition_start, len(cells)):
            token = cells[idx].strip()
            if not token:
                continue
            if condition_parts and token == condition_parts[-1]:
                continue
            condition_parts.append(token)
        if not condition_parts:
            continue
        path = logic_tokens + condition_parts
        if current_control not in by_control:
            by_control[current_control] = []
            order.append(current_control)
        by_control[current_control].append(
            TwoColumnRow(
                row_no=actual_row,
                control=current_control,
                condition_raw=" / ".join(path),
                condition_cells=path,
                indentation_level=max(len(path) - 1, 0),
                detected_type="logic_path",
                parsed_hint=" -> ".join(path),
                source={**source, "row": actual_row},
            )
        )

    blocks: list[dict[str, Any]] = []
    for idx, control_name in enumerate(order, start=1):
        table = ParsedTwoColumnTable(
            table_id=f"{table_id}_{idx:02d}",
            control_name=control_name,
            rows=by_control[control_name],
            table_kind="logic",
            source=source,
        )
        block = parse_table_to_logic_block(table)
        if block.get("parse_status") != "failed":
            blocks.append(block)
    return blocks


def _parse_group_definition_rows(
    region_rows: list[tuple[int, list[str]]],
    source: dict[str, Any],
) -> list[dict[str, Any]]:
    if len(region_rows) < 2:
        return []
    header = [str(cell or "").strip().lower() for cell in region_rows[0][1]]
    if not header or "condition" not in " ".join(header):
        return []
    control_idx = 0
    condition_idx = 1 if len(header) > 1 else 0
    definition_idx = 2 if len(header) > 2 else 1
    rows: list[dict[str, Any]] = []
    for actual_row, cells in region_rows[1:]:
        cells = [(cell or "").strip() for cell in cells]
        if control_idx >= len(cells) or condition_idx >= len(cells):
            continue
        group = cells[control_idx]
        condition = cells[condition_idx]
        definition = cells[definition_idx] if definition_idx < len(cells) else ""
        if not group or not condition:
            continue
        if condition.upper() in {"AND", "OR", "NOT"}:
            rows.append(
                {
                    "name": group,
                    "definition": f"Composite condition group ({condition.upper()})",
                    "group": group,
                    "source": {**source, "row": actual_row, "kind": "excel_group_header"},
                }
            )
            continue
        rows.append(
            {
                "name": condition,
                "definition": definition or group,
                "group": group,
                "source": {**source, "row": actual_row, "kind": "excel_group_definition"},
            }
        )
    return rows


def _parse_transition_interpretation_region(
    region_rows: list[tuple[int, list[str]]],
    source: dict[str, Any],
) -> list[dict[str, Any]]:
    if len(region_rows) < 2:
        return []
    header = [str(cell or "").strip().lower() for cell in region_rows[0][1]]
    joined = " ".join(header)
    if "transition id" not in joined or "from state" not in joined or "to state" not in joined:
        return []

    def _find_idx(name: str) -> int | None:
        for idx, token in enumerate(header):
            if name in token:
                return idx
        return None

    tid_idx = _find_idx("transition id")
    from_idx = _find_idx("from state")
    to_idx = _find_idx("to state")
    event_idx = _find_idx("trigger")
    logic_idx = _find_idx("required logic")
    timing_idx = _find_idx("timing")
    input_idx = _find_idx("input stimulus")
    output_idx = _find_idx("expected output")
    link_idx = _find_idx("source / diagram link")

    transitions: list[dict[str, Any]] = []
    for actual_row, cells in region_rows[1:]:
        cells = [(cell or "").strip() for cell in cells]
        if tid_idx is None or tid_idx >= len(cells):
            continue
        tid_value = cells[tid_idx]
        if not tid_value:
            continue
        from_state = cells[from_idx] if from_idx is not None and from_idx < len(cells) else None
        to_state = cells[to_idx] if to_idx is not None and to_idx < len(cells) else None
        event = cells[event_idx] if event_idx is not None and event_idx < len(cells) else ""
        raw_logic = cells[logic_idx] if logic_idx is not None and logic_idx < len(cells) else ""
        timing = cells[timing_idx] if timing_idx is not None and timing_idx < len(cells) else ""
        inputs = cells[input_idx] if input_idx is not None and input_idx < len(cells) else ""
        outputs = cells[output_idx] if output_idx is not None and output_idx < len(cells) else ""
        diagram_link = cells[link_idx] if link_idx is not None and link_idx < len(cells) else ""
        raw_condition = " | ".join(part for part in [raw_logic, timing, inputs, diagram_link] if part)
        transitions.append(
            {
                "id": tid_value,
                "from_state": from_state,
                "to_state": to_state,
                "event": event or tid_value,
                "raw_condition": raw_condition or event or tid_value,
                "source": {**source, "row": actual_row, "kind": "excel_transition_table"},
                "confidence": "high",
                "review_required": True,
                "derivation": "excel_transition_interpretation",
                "outputs": [outputs] if outputs else [],
                "input_stimulus": inputs,
                "timing": timing,
                "required_logic": raw_logic,
                "diagram_link": diagram_link,
            }
        )
    return transitions


def _extract_cell_diagram_hints(
    rows: list[tuple[int, list[str]]],
    source: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    diagrams: list[dict[str, Any]] = []
    state_rules: list[dict[str, Any]] = []
    seen_cells: set[str] = set()
    for actual_row, cells in rows:
        for col_idx, raw in enumerate(cells, start=1):
            text = str(raw or "").strip()
            if not text or "\n" not in text or text in seen_cells:
                continue
            seen_cells.add(text)
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            if not lines:
                continue
            title = lines[0]
            if STATE_CELL_TITLE_RE.match(title) and any("=" in line for line in lines[1:]):
                state_name = title.upper().replace(" ", "_")
                diagrams.append(
                    {
                        "file": source.get("file"),
                        "parent_document": source.get("file"),
                        "name": f"cell_state_{state_name}_{actual_row}_{col_idx}",
                        "sheet": source.get("sheet"),
                        "source_kind": "excel_cell_diagram",
                        "ocr_text": text,
                        "state_name": state_name,
                        "row": actual_row,
                        "col": col_idx,
                    }
                )
                for line in lines[1:]:
                    if "=" not in line:
                        continue
                    left, right = line.split("=", 1)
                    state_rules.append(
                        {
                            "name": left.strip().upper().replace(" ", "_"),
                            "expression": right.strip(),
                            "source": {
                                **source,
                                "row": actual_row,
                                "col": col_idx,
                                "state": state_name,
                                "kind": "excel_cell_diagram",
                            },
                        }
                    )
    return diagrams, state_rules


def extract_excel_workbook(
    path: Path,
    state_patterns: list[str],
    *,
    include_comments: bool = False,
) -> dict[str, Any]:
    """Extract logic blocks, transitions, and legacy transition rows from workbook."""
    import re as _re

    wb = load_workbook(path, data_only=True)
    out_sheets: list[dict[str, Any]] = []
    transitions: list[dict[str, Any]] = []
    logic_blocks: list[dict[str, Any]] = []
    condition_definitions: list[dict[str, Any]] = []
    merged_cell_evidence: list[dict[str, Any]] = []
    region_summaries: list[dict[str, Any]] = []
    extra_diagram_meta: list[dict[str, Any]] = []
    extra_state_rules: list[dict[str, Any]] = []
    review_annotations: list[dict[str, Any]] = []
    compiled = []
    for p in state_patterns:
        try:
            compiled.append(_re.compile(p, re.IGNORECASE))
        except _re.error:
            continue

    signals: list[dict[str, Any]] = []
    tid = 0
    file_name = path.name
    selected_sheets = _sheet_names_for_workbook(list(wb.sheetnames))
    for sheet_name in selected_sheets:
        ws = wb[sheet_name]
        merged_cell_evidence.extend(collect_merged_cell_evidence(ws, file_name, sheet_name))
        if include_comments:
            review_annotations.extend(collect_sheet_comments(ws, file_name, sheet_name))
        sheet_rows = _sheet_to_filled_rows(ws)
        grid = [cells for _, cells in sheet_rows]
        source = {"file": path.name, "sheet": sheet_name}
        signals.extend(parse_signal_grid(grid, source))
        cell_diagrams, cell_state_rules = _extract_cell_diagram_hints(sheet_rows, source)
        extra_diagram_meta.extend(cell_diagrams)
        extra_state_rules.extend(cell_state_rules)
        regions = _find_nonempty_regions(sheet_rows)
        sections = _section_slices(sheet_rows)
        out_sheets.append(
            {
                "name": sheet_name,
                "row_count": len(grid),
                "sample_rows": grid[:30],
                "regions": [region["bbox"] for region in regions],
                "sections": [section["title"] for section in sections],
            }
        )
        region_summaries.extend(
            [{**region["bbox"], "sheet": sheet_name, "nonempty_cells": region["nonempty_cells"]} for region in regions]
        )

        for section_index, section in enumerate(sections, start=1):
            title = str(section["title"] or "")
            section_rows = section["rows"][1:]
            if len(section_rows) < 2:
                continue
            section_source = {**source, "section": title, "section_index": section_index}
            if "control conditions" in title.lower() or "condition group definitions" in title.lower():
                logic_blocks.extend(
                    _parse_event_condition_logic_region(
                        section_rows,
                        section_source,
                        table_id=f"XL_{sheet_name[:8]}_SEC_{section_index:02d}",
                    )
                )
                condition_definitions.extend(_parse_group_definition_rows(section_rows, section_source))
            if "state transition interpretation" in title.lower():
                transitions.extend(_parse_transition_interpretation_region(section_rows, section_source))

        parse_regions = regions or [{"rows": sheet_rows, "bbox": None, "nonempty_cells": sum(1 for row in grid for cell in row if cell)}]
        for region_index, region in enumerate(parse_regions, start=1):
            region_rows = region["rows"]
            region_grid = [cells for _, cells in region_rows]
            region_row_numbers = [row_no for row_no, _ in region_rows]
            if len(region_grid) < 2:
                continue
            header_idx = _find_region_header_index(region_grid)
            parse_grid = region_grid[header_idx:]
            parse_row_numbers = region_row_numbers[header_idx:]
            if len(parse_grid) < 2:
                continue
            hdr = [c.lower() for c in parse_grid[0]]
            hdr_joined = " ".join(hdr)
            region_source = {
                **source,
                "region": region_index,
                **({"bbox": region["bbox"]} if region.get("bbox") else {}),
            }

            if hdr and "condition" in hdr[0] and len(hdr) > 1 and "definition" in hdr[1]:
                for ri, cells in zip(parse_row_numbers[1:], parse_grid[1:]):
                    if len(cells) < 2 or not cells[0]:
                        continue
                    condition_definitions.append(
                        {
                            "name": cells[0].strip(),
                            "definition": cells[1].strip(),
                            "source": {**source, "row": ri},
                        }
                    )
                continue
            if hdr and ("constant" in hdr[0] or "parameter" in hdr[0] or "name" in hdr[0]) and len(hdr) > 1 and ("value" in hdr[1] or "definition" in hdr[1]):
                for ri, cells in zip(parse_row_numbers[1:], parse_grid[1:]):
                    if len(cells) < 2 or not cells[0]:
                        continue
                    condition_definitions.append(
                        {
                            "name": cells[0].strip(),
                            "definition": cells[1].strip(),
                            "source": {**source, "row": ri},
                        }
                    )
                continue
            if _looks_like_event_condition_region(parse_grid[0]):
                logic_blocks.extend(
                    _parse_event_condition_logic_region(
                        list(zip(parse_row_numbers, parse_grid)),
                        region_source,
                        table_id=f"XL_{sheet_name[:8]}_{region_index:02d}",
                    )
                )
                continue
            if "control" in hdr_joined and "condition" in hdr_joined and "logic" not in hdr:
                for pt in parse_control_condition_grid(
                    parse_grid, region_source, table_id=f"XL_{sheet_name[:8]}"
                ):
                    if pt.table_kind == "logic":
                        block = parse_table_to_logic_block(pt)
                        logic_blocks.append(block)
                    elif pt.table_kind == "constant":
                        for row in pt.rows:
                            condition_definitions.append(
                                {
                                    "name": row.control,
                                    "definition": row.condition_raw,
                                    "source": row.source,
                                    "constant_value": row.parsed_hint,
                                }
                            )
                continue
            if "logic" in hdr and "condition" in hdr:
                for block in rows_from_grid(
                    parse_grid,
                    region_source,
                    block_id_prefix=f"XL_{sheet_name[:8]}",
                    row_numbers=parse_row_numbers,
                    preserve_layout=True,
                ):
                    logic_blocks.append(
                        {
                            "id": block.id,
                            "name": block.name,
                            "raw_expression": block.raw_expression,
                            "tree": block.tree,
                            "block_type": block.block_type,
                            "source": block.source,
                            "review_required": True,
                        }
                    )
                    tid += 1
                    transitions.append(
                        {
                            "id": f"TR_{tid:03d}",
                            "from_state": None,
                            "to_state": None,
                            "event": block.name,
                            "raw_condition": block.raw_expression,
                            "condition_tree": block.tree,
                            "source": {**source, "logic_block_id": block.id},
                            "confidence": "medium" if block.parse_status != "failed" else "low",
                            "review_required": True,
                        }
                    )
                continue

            if "previous" in " ".join(hdr) or "transition" in " ".join(hdr):
                transitions.extend(
                    parse_transition_table(
                        parse_grid,
                        region_source,
                        row_numbers=parse_row_numbers,
                        preserve_layout=True,
                    )
                )
                continue

            # Fallback: ADM state pattern rows
            for actual_row, cells in region_rows:
                joined = " ".join(cells)
                state_hits = sum(1 for c in cells if any(rx.search(c) for rx in compiled))
                if state_hits >= 1 and len([c for c in cells if c]) >= 2:
                    tid += 1
                    transitions.append(
                        {
                            "id": f"TR_{tid:03d}",
                            "from_state": cells[0] if cells else None,
                            "to_state": cells[1] if len(cells) > 1 else None,
                            "event": cells[2] if len(cells) > 2 else None,
                            "raw_condition": cells[3] if len(cells) > 3 and cells[3] else joined,
                            "source": {**region_source, "row_hint": actual_row},
                            "confidence": "low",
                            "review_required": True,
                        }
                    )

    all_sheet_names = list(wb.sheetnames)
    wb.close()
    sheet_scan_summary = [
        {
            "name": name,
            "selected": name in selected_sheets,
            "logic_blocks": sum(
                1 for b in logic_blocks if str((b.get("source") or {}).get("sheet") or "") == name
            ),
        }
        for name in all_sheet_names
    ]
    drawing = extract_excel_drawing_semantics(path)
    return {
        "file": str(path),
        "sheets": out_sheets,
        "sheet_regions": region_summaries,
        "sheet_scan_summary": sheet_scan_summary,
        "sheets_selected": selected_sheets,
        "logic_blocks": logic_blocks,
        "signals": signals,
        "condition_definitions": condition_definitions,
        "transition_candidates": transitions + drawing.get("transition_candidates", []),
        "drawing_shapes": drawing.get("drawing_shapes", []),
        "drawing_connectors": drawing.get("drawing_connectors", []),
        "diagram_meta": extra_diagram_meta + drawing.get("diagram_meta", []),
        "state_rules": extra_state_rules + drawing.get("state_rules", []),
        "merged_cell_evidence": merged_cell_evidence,
        "review_annotations": review_annotations,
    }

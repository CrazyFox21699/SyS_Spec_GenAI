"""Job diagnostic helpers for parser and Copilot troubleshooting."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from src.utils.yaml_utils import load_yaml
from web.bundle_store import load_split_bundle


def _word_table_header_kind(header: list[str]) -> str:
    joined = " ".join(str(h or "").lower() for h in header)
    if "logic" in joined and "condition" in joined:
        return "logic_condition"
    control_tokens = (
        "control",
        "event",
        "function",
        "item",
        "judgment",
        "signal",
        "permission",
        "prohibition",
    )
    has_control = any(t in joined for t in control_tokens)
    has_condition = "condition" in joined or " cond" in joined
    if has_control and has_condition:
        return "two_column_logic"
    if "given" in joined and ("expected" in joined or "when" in joined):
        return "test_reference"
    if "previous" in joined or "next" in joined:
        return "transition"
    if header and "condition" in str(header[0]).lower():
        return "condition_definition"
    return "unmatched"


def diagnose_word_tables(classified_files: list[dict[str, Any]], uploads_dir: Path | None = None) -> dict[str, Any]:
    from docx import Document

    docx_files = [
        row for row in classified_files if str(row.get("file_type") or "").lower() in {"word", "docx"}
    ]
    if not docx_files and uploads_dir:
        docx_files = [{"file": str(p), "name": p.name} for p in uploads_dir.glob("*.docx")]

    tables_total = 0
    matched = 0
    unmatched = 0
    by_kind: dict[str, int] = {}
    samples: list[dict[str, Any]] = []

    for row in docx_files[:5]:
        path = Path(str(row.get("file") or row.get("name") or ""))
        if not path.is_file():
            if uploads_dir:
                alt = uploads_dir / path.name
                if alt.is_file():
                    path = alt
            if not path.is_file():
                continue
        try:
            doc = Document(path)
        except Exception as exc:  # noqa: BLE001
            samples.append({"file": path.name, "error": str(exc)})
            continue
        for ti, table in enumerate(doc.tables):
            grid = [[c.text.strip() for c in tr.cells] for tr in table.rows]
            if not grid:
                continue
            tables_total += 1
            header = [c.lower() for c in grid[0]]
            kind = _word_table_header_kind(header)
            by_kind[kind] = by_kind.get(kind, 0) + 1
            if kind == "unmatched":
                unmatched += 1
                if len(samples) < 8:
                    samples.append({"file": path.name, "table": ti + 1, "header": grid[0][:6]})
            else:
                matched += 1

    return {
        "docx_files_scanned": len(docx_files),
        "tables_total": tables_total,
        "tables_matched": matched,
        "tables_unmatched": unmatched,
        "by_kind": by_kind,
        "unmatched_samples": samples,
    }


def diagnose_excel_sheets(classified_files: list[dict[str, Any]], bundle: dict[str, Any]) -> dict[str, Any]:
    excel_import = bundle.get("excel_import") or {}
    if excel_import.get("sheets"):
        return {"source": "import", "sheets": excel_import["sheets"]}

    sheets: list[dict[str, Any]] = []
    for row in classified_files:
        if str(row.get("file_type") or "").lower() not in {"excel", "xlsx", "xlsm"}:
            continue
        path = Path(str(row.get("file") or ""))
        if not path.is_file():
            continue
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            for name in wb.sheetnames:
                sheets.append({"name": name, "source_file": path.name})
        except Exception as exc:  # noqa: BLE001
            sheets.append({"name": "?", "source_file": path.name, "error": str(exc)})

    summary = bundle.get("summary") or {}
    if summary.get("excel_sheets"):
        return {"source": "analyze", "sheets": summary["excel_sheets"], "discovered": sheets}
    return {"source": "classified_files", "sheets": sheets}


def diagnose_logic_blocks(bundle: dict[str, Any]) -> dict[str, Any]:
    blocks = bundle.get("logic_blocks") or []
    resolved = bundle.get("resolved_logic_blocks") or []
    by_status: dict[str, int] = {}
    items: list[dict[str, Any]] = []
    for block in blocks[:50]:
        status = str(block.get("parse_status") or "unknown")
        by_status[status] = by_status.get(status, 0) + 1
        items.append(
            {
                "id": block.get("id"),
                "name": block.get("name"),
                "parse_status": status,
                "gate_status": block.get("gate_status"),
            }
        )
    return {
        "count": len(blocks),
        "resolved_count": len(resolved),
        "by_parse_status": by_status,
        "items": items,
        "bootstrap_source": bundle.get("bootstrap_source"),
    }


def diagnose_job_bundle(bundle: dict[str, Any], *, uploads_dir: Path | None = None) -> dict[str, Any]:
    classified = bundle.get("classified_files") or []
    tc_count = len(bundle.get("test_candidates") or [])
    overlay_count = len((bundle.get("ai_assists") or {}).get("candidate_overlays") or {})
    return {
        "bootstrap_source": bundle.get("bootstrap_source") or "analyze",
        "bootstrap_label": bundle.get("bootstrap_label") or "",
        "test_candidates": tc_count,
        "candidate_overlays": overlay_count,
        "logic": diagnose_logic_blocks(bundle),
        "word_tables": diagnose_word_tables(classified, uploads_dir=uploads_dir),
        "excel_sheets": diagnose_excel_sheets(classified, bundle),
        "summary": bundle.get("summary") or {},
    }


def load_bundle_for_diagnostic(job_output_dir: Path) -> dict[str, Any] | None:
    split = load_split_bundle(job_output_dir)
    if split:
        return split
    legacy = job_output_dir / "ui_bundle.yaml"
    if legacy.exists():
        return load_yaml(legacy)
    return None
